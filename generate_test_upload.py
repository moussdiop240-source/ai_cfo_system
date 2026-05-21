"""Generate a test Excel upload file on the Desktop."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_UPLOAD_FIELDS = [
    ("INCOME STATEMENT",  "revenue",                    15_000_000, "Total net revenue"),
    ("INCOME STATEMENT",  "cogs",                        4_500_000, "Cost of goods sold"),
    ("INCOME STATEMENT",  "gross_profit",               10_500_000, "Revenue − COGS"),
    ("INCOME STATEMENT",  "operating_expenses",          7_000_000, "Total operating expenses"),
    ("INCOME STATEMENT",  "rd_expense",                  2_800_000, "Research & development"),
    ("INCOME STATEMENT",  "sg_a_expense",                4_200_000, "SG&A expense"),
    ("INCOME STATEMENT",  "ebitda",                      3_500_000, "EBITDA"),
    ("INCOME STATEMENT",  "depreciation",                  350_000, "Depreciation & amortization"),
    ("INCOME STATEMENT",  "ebit",                        3_150_000, "EBIT"),
    ("INCOME STATEMENT",  "interest_expense",              180_000, "Net interest expense"),
    ("INCOME STATEMENT",  "pre_tax_income",              2_970_000, "Pre-tax income (EBT)"),
    ("INCOME STATEMENT",  "tax_provision",                 654_000, "Income tax provision"),
    ("INCOME STATEMENT",  "net_income",                  2_316_000, "Net income"),
    ("BALANCE SHEET",     "total_assets",               65_000_000, "Total assets"),
    ("BALANCE SHEET",     "current_assets",             28_000_000, "Current assets"),
    ("BALANCE SHEET",     "cash",                       13_000_000, "Cash & cash equivalents"),
    ("BALANCE SHEET",     "accounts_receivable",        10_000_000, "Net accounts receivable"),
    ("BALANCE SHEET",     "inventory",                     300_000, "Inventory"),
    ("BALANCE SHEET",     "prepaid_expenses",            4_700_000, "Prepaid expenses"),
    ("BALANCE SHEET",     "total_equity",               39_000_000, "Total stockholders equity"),
    ("BALANCE SHEET",     "current_liabilities",        11_000_000, "Current liabilities"),
    ("BALANCE SHEET",     "accounts_payable",            3_500_000, "Accounts payable"),
    ("BALANCE SHEET",     "deferred_revenue",            5_000_000, "Deferred revenue"),
    ("BALANCE SHEET",     "total_debt",                 15_000_000, "Total debt"),
    ("BALANCE SHEET",     "long_term_debt",             12_500_000, "Long-term debt"),
    ("BALANCE SHEET",     "goodwill",                   10_000_000, "Goodwill (ASC 350 / IAS 36)"),
    ("BALANCE SHEET",     "rou_assets",                  5_200_000, "Right-of-use assets (ASC 842)"),
    ("BALANCE SHEET",     "lease_liability",             5_000_000, "Lease liability (ASC 842)"),
    ("BALANCE SHEET",     "allowance_for_credit_losses",   500_000, "ACL / ECL (CECL / IFRS 9)"),
    ("CASH FLOW",         "cash_from_operations",        4_200_000, "Net cash from operations"),
    ("CASH FLOW",         "capex",                         900_000, "Capital expenditures"),
    ("CASH FLOW",         "free_cash_flow",              3_300_000, "FCF = Operating CF - CapEx"),
    ("CASH FLOW",         "monthly_cash_burn",                   0, "Monthly cash burn (0 if profitable)"),
    ("CASH FLOW",         "operating_lease_expense",       420_000, "Operating lease expense"),
    ("EQUITY",            "shares_outstanding",          9_000_000, "Basic shares outstanding"),
    ("EQUITY",            "diluted_shares",              9_500_000, "Diluted shares"),
    ("SAAS METRICS",      "arr",                        60_000_000, "Annual recurring revenue"),
    ("SAAS METRICS",      "nrr_pct",                           122, "Net revenue retention %"),
    ("SAAS METRICS",      "churn_rate_pct",                    3.8, "Gross churn rate %"),
    ("SAAS METRICS",      "headcount",                         245, "Total full-time employees"),
    ("BUDGET VS ACTUALS", "actuals_revenue",            15_000_000, "Actual revenue"),
    ("BUDGET VS ACTUALS", "actuals_cogs",                4_500_000, "Actual COGS"),
    ("BUDGET VS ACTUALS", "actuals_gross_profit",       10_500_000, "Actual gross profit"),
    ("BUDGET VS ACTUALS", "actuals_ebitda",              3_500_000, "Actual EBITDA"),
    ("BUDGET VS ACTUALS", "actuals_rd_expense",          2_800_000, "Actual R&D"),
    ("BUDGET VS ACTUALS", "actuals_sg_a",                4_200_000, "Actual SG&A"),
    ("BUDGET VS ACTUALS", "budget_revenue",             14_000_000, "Budget revenue"),
    ("BUDGET VS ACTUALS", "budget_cogs",                 4_200_000, "Budget COGS"),
    ("BUDGET VS ACTUALS", "budget_gross_profit",         9_800_000, "Budget gross profit"),
    ("BUDGET VS ACTUALS", "budget_ebitda",               3_100_000, "Budget EBITDA"),
    ("BUDGET VS ACTUALS", "budget_rd_expense",           2_500_000, "Budget R&D"),
    ("BUDGET VS ACTUALS", "budget_sg_a",                 3_900_000, "Budget SG&A"),
    ("HISTORICAL REVENUE","hist_q1",                     8_500_000, "Quarter 1 (oldest)"),
    ("HISTORICAL REVENUE","hist_q2",                     9_200_000, "Quarter 2"),
    ("HISTORICAL REVENUE","hist_q3",                    10_100_000, "Quarter 3"),
    ("HISTORICAL REVENUE","hist_q4",                    10_800_000, "Quarter 4"),
    ("HISTORICAL REVENUE","hist_q5",                    11_500_000, "Quarter 5"),
    ("HISTORICAL REVENUE","hist_q6",                    12_300_000, "Quarter 6"),
    ("HISTORICAL REVENUE","hist_q7",                    13_200_000, "Quarter 7"),
    ("HISTORICAL REVENUE","hist_q8",                    14_100_000, "Quarter 8"),
    ("HISTORICAL REVENUE","hist_q9",                    15_000_000, "Quarter 9 (most recent)"),
]

SECTION_FILL = PatternFill("solid", fgColor="1E293B")
INPUT_FILL   = PatternFill("solid", fgColor="EFF6FF")


def make_template(fields):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 44

    ws.merge_cells("A1:C1")
    ws["A1"].value = "AI CFO System - Financial Data Template"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="060D1F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2")
    ws["A2"].value = "Instructions: fill Column B (blue cells). Do not edit Column A or C."
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    for col, text in [(1, "Field"), (2, "Value"), (3, "Description")]:
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E40AF")
        cell.alignment = Alignment(horizontal="center")

    row, prev_sec = 4, None
    for section, key, default, desc in fields:
        if section != prev_sec:
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1, value=section)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = SECTION_FILL
            ws.row_dimensions[row].height = 18
            row += 1
            prev_sec = section

        ws.cell(row=row, column=1, value=key).font = Font(size=10)
        val = ws.cell(row=row, column=2, value=default)
        val.font = Font(color="1D4ED8", size=10)
        val.fill = INPUT_FILL
        val.alignment = Alignment(horizontal="right")
        if isinstance(default, float) and default < 100:
            val.number_format = "0.0"
        elif isinstance(default, (int, float)) and default >= 1000:
            val.number_format = "#,##0"
        ws.cell(row=row, column=3, value=desc).font = Font(color="6B7280", size=9)
        row += 1

    return wb


if __name__ == "__main__":
    wb = make_template(_UPLOAD_FIELDS)
    path = os.path.join(os.path.expanduser("~"), "Desktop", "cfo_test_upload.xlsx")
    wb.save(path)
    print(f"Saved: {path}")
