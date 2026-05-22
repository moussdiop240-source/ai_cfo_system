"""
AI CFO Multi-Agent System — Streamlit UI
Supports Anthropic (API key) and Ollama (local, no key needed).
Run: streamlit run streamlit_app.py
"""
import os
import sys

sys.path.insert(0, ".")

import pandas as pd
import streamlit as st

# ── page config (must be first st call) ─────────────────────────────────────
st.set_page_config(
    page_title="AI CFO System",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
  [data-testid="stSidebar"] { background: #0B1120; }
  .metric-card { background: #0B1120; border: 1px solid rgba(255,255,255,0.08);
                 border-radius: 10px; padding: 16px; margin-bottom: 8px; }
  .compliant  { color: #10B981; font-weight: 700; }
  .disclosure { color: #FBBF24; font-weight: 700; }
  .noncompliant{ color: #F87171; font-weight: 700; }
  .round-header { background: linear-gradient(135deg,#0B1120,#141e33);
                  border-left: 4px solid; padding: 12px 16px;
                  border-radius: 0 8px 8px 0; margin-bottom: 12px; }
  .stButton button { width: 100%; }
</style>
""", unsafe_allow_html=True)

# ── default synthetic financial data ─────────────────────────────────────────
DEFAULT_DATA = {
    "revenue": 12_840_000, "cogs": 3_594_000, "gross_profit": 9_246_000,
    "operating_expenses": 6_120_000, "rd_expense": 2_430_000, "sg_a_expense": 3_690_000,
    "ebitda": 3_126_000, "depreciation": 312_000, "ebit": 2_814_000,
    "interest_expense": 210_000, "pre_tax_income": 2_604_000,
    "tax_provision": 573_000, "net_income": 2_031_000,
    "total_assets": 58_400_000, "current_assets": 24_100_000, "cash": 11_250_000,
    "accounts_receivable": 8_640_000, "inventory": 210_000, "prepaid_expenses": 4_000_000,
    "total_equity": 34_200_000, "current_liabilities": 9_800_000,
    "accounts_payable": 3_150_000, "deferred_revenue": 4_200_000,
    "total_debt": 14_000_000, "long_term_debt": 11_800_000,
    "cash_from_operations": 3_840_000, "capex": 780_000, "free_cash_flow": 3_060_000,
    "monthly_cash_burn": 0,
    "shares_outstanding": 8_200_000, "diluted_shares": 8_650_000,
    "rou_assets": 4_800_000, "lease_liability": 4_620_000,
    "operating_lease_expense": 360_000,
    "goodwill": 9_600_000, "goodwill_impairment_test_date": "2026-01-31",
    "impairment_test_performed": True, "impairment_tested_at_cgu_level": True,
    "allowance_for_credit_losses": 432_000,
    "ecl_stage1_allowance": 258_000, "ecl_stage2_allowance": 129_000, "ecl_stage3_allowance": 45_000,
    "revenue_recognition_policy": "ASC 606 5-step model",
    "inventory_cost_method": "fifo",
    "interest_cash_flow_classification": "operating",
    "cash_flow_policy_consistent": True, "comparative_period_presented": True,
    "publicly_listed": True, "qualifying_development_projects": True,
    "rd_dev_capitalizable_pct": 0.35,
    "actuals": {"revenue": 12_840_000, "cogs": 3_594_000, "gross_profit": 9_246_000,
                "ebitda": 3_126_000, "rd_expense": 2_430_000, "sg_a": 3_690_000},
    "budget":  {"revenue": 11_500_000, "cogs": 3_335_000, "gross_profit": 8_165_000,
                "ebitda": 2_700_000, "rd_expense": 2_100_000, "sg_a": 3_365_000},
    "historical_revenue": [7_200_000, 7_810_000, 8_450_000, 9_120_000,
                           9_980_000, 10_620_000, 11_310_000, 11_870_000, 12_840_000],
    "segments": [
        {"name": "Enterprise",           "revenue": 7_704_000, "gross_profit": 5_852_400, "assets": 32_000_000},
        {"name": "SMB",                  "revenue": 3_852_000, "gross_profit": 2_813_160, "assets": 16_000_000},
        {"name": "Professional Services","revenue": 1_284_000, "gross_profit":   580_440, "assets":  6_000_000},
    ],
    "arr": 51_360_000, "nrr_pct": 118, "churn_rate_pct": 4.2,
    "headcount": 214,
}

# ── helpers ───────────────────────────────────────────────────────────────────
def fmt(v, prefix="$", suffix=""):
    if abs(v) >= 1_000_000: return f"{prefix}{v/1_000_000:.1f}M{suffix}"
    if abs(v) >= 1_000:     return f"{prefix}{v/1_000:.0f}K{suffix}"
    return f"{prefix}{v:,.0f}{suffix}"

def status_badge(s):
    m = {"COMPLIANT": "compliant", "DISCLOSURE_REQUIRED": "disclosure", "NON_COMPLIANT": "noncompliant"}
    icons = {"COMPLIANT": "✓", "DISCLOSURE_REQUIRED": "⚠", "NON_COMPLIANT": "✗"}
    cls = m.get(s, "disclosure")
    return f'<span class="{cls}">{icons.get(s,"?")} {s}</span>'

def run_deterministic(data, company, period):
    from backend.agents.math_engine import FinancialCalculationEngine
    from backend.compliance.gaap_engine import GAAPEngine
    from backend.compliance.ifrs_engine import IFRSEngine
    from backend.rag.pipeline import RAGPipeline

    eng  = FinancialCalculationEngine()
    kpis = eng.calculate_kpis(data)
    var  = eng.calculate_variance_analysis(data["actuals"], data["budget"])
    anom = eng.detect_anomalies(data, kpis)
    run  = eng.calculate_cash_runway(data, kpis)
    fcast= eng.forecast_revenue(data["historical_revenue"], periods=8)
    recon= eng.calculate_reconciliation(data, data)

    gaap_e = GAAPEngine()
    ifrs_e = IFRSEngine()
    gaap   = gaap_e.check_all(data, kpis, var, run)
    ifrs   = ifrs_e.check_all(data, kpis, var, run)

    rag_pipeline = RAGPipeline()
    rag_q = rag_pipeline.build_rag_query({
        "task_description": f"{period} board analysis {company}",
        "task_type": "full_report", "period": period,
        "kpi_metrics": kpis, "anomaly_flags": anom,
        "gaap_results": gaap, "ifrs_results": ifrs,
    })
    chunks = rag_pipeline.retrieve(rag_q, top_k=5)

    return dict(
        kpis=kpis, variance=var, anomalies=anom, runway=run,
        forecast=fcast, reconcile=recon,
        gaap=gaap, ifrs=ifrs,
        rag_chunks=[c.to_dict() for c in chunks], rag_query=rag_q,
        company=company, period=period, data=data,
    )


# ── Excel upload helpers ──────────────────────────────────────────────────────
_UPLOAD_FIELDS = [
    ("INCOME STATEMENT",  "revenue",                    12_840_000, "Total net revenue"),
    ("INCOME STATEMENT",  "cogs",                        3_594_000, "Cost of goods sold"),
    ("INCOME STATEMENT",  "gross_profit",                9_246_000, "Revenue − COGS"),
    ("INCOME STATEMENT",  "operating_expenses",          6_120_000, "Total operating expenses"),
    ("INCOME STATEMENT",  "rd_expense",                  2_430_000, "Research & development"),
    ("INCOME STATEMENT",  "sg_a_expense",                3_690_000, "SG&A expense"),
    ("INCOME STATEMENT",  "ebitda",                      3_126_000, "EBITDA"),
    ("INCOME STATEMENT",  "depreciation",                  312_000, "Depreciation & amortization"),
    ("INCOME STATEMENT",  "ebit",                        2_814_000, "EBIT"),
    ("INCOME STATEMENT",  "interest_expense",              210_000, "Net interest expense"),
    ("INCOME STATEMENT",  "pre_tax_income",              2_604_000, "Pre-tax income (EBT)"),
    ("INCOME STATEMENT",  "tax_provision",                 573_000, "Income tax provision"),
    ("INCOME STATEMENT",  "net_income",                  2_031_000, "Net income"),
    ("BALANCE SHEET",     "total_assets",               58_400_000, "Total assets"),
    ("BALANCE SHEET",     "current_assets",             24_100_000, "Current assets"),
    ("BALANCE SHEET",     "cash",                       11_250_000, "Cash & cash equivalents"),
    ("BALANCE SHEET",     "accounts_receivable",         8_640_000, "Net accounts receivable"),
    ("BALANCE SHEET",     "inventory",                     210_000, "Inventory"),
    ("BALANCE SHEET",     "prepaid_expenses",            4_000_000, "Prepaid expenses"),
    ("BALANCE SHEET",     "total_equity",               34_200_000, "Total stockholders equity"),
    ("BALANCE SHEET",     "current_liabilities",         9_800_000, "Current liabilities"),
    ("BALANCE SHEET",     "accounts_payable",            3_150_000, "Accounts payable"),
    ("BALANCE SHEET",     "deferred_revenue",            4_200_000, "Deferred revenue"),
    ("BALANCE SHEET",     "total_debt",                 14_000_000, "Total debt"),
    ("BALANCE SHEET",     "long_term_debt",             11_800_000, "Long-term debt"),
    ("BALANCE SHEET",     "goodwill",                    9_600_000, "Goodwill (ASC 350 / IAS 36)"),
    ("BALANCE SHEET",     "rou_assets",                  4_800_000, "Right-of-use assets (ASC 842)"),
    ("BALANCE SHEET",     "lease_liability",             4_620_000, "Lease liability (ASC 842)"),
    ("BALANCE SHEET",     "allowance_for_credit_losses",   432_000, "ACL / ECL (CECL / IFRS 9)"),
    ("CASH FLOW",         "cash_from_operations",        3_840_000, "Net cash from operations"),
    ("CASH FLOW",         "capex",                         780_000, "Capital expenditures"),
    ("CASH FLOW",         "free_cash_flow",              3_060_000, "FCF = Operating CF − CapEx"),
    ("CASH FLOW",         "monthly_cash_burn",                   0, "Monthly cash burn (0 if profitable)"),
    ("CASH FLOW",         "operating_lease_expense",       360_000, "Operating lease expense"),
    ("EQUITY",            "shares_outstanding",          8_200_000, "Basic shares outstanding"),
    ("EQUITY",            "diluted_shares",              8_650_000, "Diluted shares"),
    ("SAAS METRICS",      "arr",                        51_360_000, "Annual recurring revenue (0 if N/A)"),
    ("SAAS METRICS",      "nrr_pct",                           118, "Net revenue retention %"),
    ("SAAS METRICS",      "churn_rate_pct",                    4.2, "Gross churn rate %"),
    ("SAAS METRICS",      "headcount",                         214, "Total full-time employees"),
    ("BUDGET VS ACTUALS", "actuals_revenue",            12_840_000, "Actual revenue"),
    ("BUDGET VS ACTUALS", "actuals_cogs",                3_594_000, "Actual COGS"),
    ("BUDGET VS ACTUALS", "actuals_gross_profit",        9_246_000, "Actual gross profit"),
    ("BUDGET VS ACTUALS", "actuals_ebitda",              3_126_000, "Actual EBITDA"),
    ("BUDGET VS ACTUALS", "actuals_rd_expense",          2_430_000, "Actual R&D"),
    ("BUDGET VS ACTUALS", "actuals_sg_a",                3_690_000, "Actual SG&A"),
    ("BUDGET VS ACTUALS", "budget_revenue",             11_500_000, "Budget revenue"),
    ("BUDGET VS ACTUALS", "budget_cogs",                 3_335_000, "Budget COGS"),
    ("BUDGET VS ACTUALS", "budget_gross_profit",         8_165_000, "Budget gross profit"),
    ("BUDGET VS ACTUALS", "budget_ebitda",               2_700_000, "Budget EBITDA"),
    ("BUDGET VS ACTUALS", "budget_rd_expense",           2_100_000, "Budget R&D"),
    ("BUDGET VS ACTUALS", "budget_sg_a",                 3_365_000, "Budget SG&A"),
    ("HISTORICAL REVENUE","hist_q1",                     7_200_000, "Quarter 1 (oldest)"),
    ("HISTORICAL REVENUE","hist_q2",                     7_810_000, "Quarter 2"),
    ("HISTORICAL REVENUE","hist_q3",                     8_450_000, "Quarter 3"),
    ("HISTORICAL REVENUE","hist_q4",                     9_120_000, "Quarter 4"),
    ("HISTORICAL REVENUE","hist_q5",                     9_980_000, "Quarter 5"),
    ("HISTORICAL REVENUE","hist_q6",                    10_620_000, "Quarter 6"),
    ("HISTORICAL REVENUE","hist_q7",                    11_310_000, "Quarter 7"),
    ("HISTORICAL REVENUE","hist_q8",                    11_870_000, "Quarter 8"),
    ("HISTORICAL REVENUE","hist_q9",                    12_840_000, "Quarter 9 (most recent)"),
]
_UPLOAD_KEYS = {row[1] for row in _UPLOAD_FIELDS}


def _make_template() -> bytes:
    import io as _io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 44

    ws.merge_cells("A1:C1")
    ws["A1"].value = "AI CFO System — Financial Data Template"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="060D1F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2")
    ws["A2"].value = "Instructions: fill Column B (blue cells). Do not edit Column A or C."
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    for col, text in [(1, "Field"), (2, "Value"), (3, "Description")]:
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E40AF")
        cell.alignment = Alignment(horizontal="center")

    SECTION_FILL = PatternFill("solid", fgColor="1E293B")
    INPUT_FILL   = PatternFill("solid", fgColor="EFF6FF")
    row, prev_sec = 4, None
    for section, key, default, desc in _UPLOAD_FIELDS:
        if section != prev_sec:
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1, value=section)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = SECTION_FILL
            ws.row_dimensions[row].height = 18
            row += 1
            prev_sec = section

        ws.cell(row=row, column=1, value=key).font = Font(size=10)
        val = ws.cell(row=row, column=2, value=default)
        val.font = Font(color="1D4ED8", size=10)
        val.fill = INPUT_FILL
        val.alignment = Alignment(horizontal="right")
        if isinstance(default, float) and default < 100:
            val.number_format = "0.0"
        elif isinstance(default, (int, float)) and default >= 1000:
            val.number_format = "#,##0"
        ws.cell(row=row, column=3, value=desc).font = Font(color="6B7280", size=9)
        row += 1

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_upload(file_bytes: bytes):
    import copy
    import io as _io

    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    flat = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""
        val = row[1]
        if key in _UPLOAD_KEYS and val is not None:
            try:
                flat[key] = float(val)
            except (TypeError, ValueError):
                flat[key] = val

    data = copy.deepcopy(DEFAULT_DATA)

    direct_keys = [
        "revenue","cogs","gross_profit","operating_expenses","rd_expense","sg_a_expense",
        "ebitda","depreciation","ebit","interest_expense","pre_tax_income","tax_provision",
        "net_income","total_assets","current_assets","cash","accounts_receivable","inventory",
        "prepaid_expenses","total_equity","current_liabilities","accounts_payable",
        "deferred_revenue","total_debt","long_term_debt","goodwill","rou_assets",
        "lease_liability","allowance_for_credit_losses","cash_from_operations","capex",
        "free_cash_flow","monthly_cash_burn","operating_lease_expense",
        "shares_outstanding","diluted_shares","arr","nrr_pct","churn_rate_pct","headcount",
    ]
    for k in direct_keys:
        if k in flat:
            data[k] = flat[k]

    actuals = {fld: flat[f"actuals_{fld}"] for fld in ["revenue","cogs","gross_profit","ebitda","rd_expense","sg_a"] if f"actuals_{fld}" in flat}
    if actuals:
        data["actuals"] = actuals

    budget = {fld: flat[f"budget_{fld}"] for fld in ["revenue","cogs","gross_profit","ebitda","rd_expense","sg_a"] if f"budget_{fld}" in flat}
    if budget:
        data["budget"] = budget

    hist = [flat[f"hist_q{i}"] for i in range(1, 10) if f"hist_q{i}" in flat]
    if hist:
        data["historical_revenue"] = hist

    return data, len(flat)


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📊 AI CFO System")
    st.caption("Multi-Agent · Anti-Hallucination · GAAP + IFRS")
    st.divider()

    # ── LLM backend ──────────────────────────────────────────────────────────
    st.markdown("### LLM Backend")
    backend_choice = st.radio(
        "Select backend",
        ["Auto-detect", "Anthropic (API key)", "Ollama (local, no key)"],
        index=0,
        label_visibility="collapsed",
    )
    backend_map = {
        "Auto-detect": "auto",
        "Anthropic (API key)": "anthropic",
        "Ollama (local, no key)": "ollama",
    }
    selected_backend = backend_map[backend_choice]

    # Anthropic key input
    if selected_backend in ("anthropic", "auto"):
        api_key_input = st.text_input(
            "Anthropic API Key",
            value=os.getenv("ANTHROPIC_API_KEY", ""),
            type="password",
            placeholder="sk-ant-...",
            help="Leave blank to use Ollama fallback",
        )
        if api_key_input:
            os.environ["ANTHROPIC_API_KEY"] = api_key_input
            from backend.llm.adapter import reset_adapter
            reset_adapter()

    # Ollama config
    if selected_backend in ("ollama", "auto"):
        ollama_host = st.text_input(
            "Ollama Host",
            value=os.getenv("OLLAMA_HOST", "http://localhost:11434"),
        )
        os.environ["OLLAMA_HOST"] = ollama_host

        # Health check + model list
        try:
            from backend.llm.adapter import LLMAdapter
            _probe = LLMAdapter(backend="ollama", ollama_host=ollama_host)
            if _probe.check_ollama_health():
                models = _probe.list_ollama_models()
                if models:
                    ollama_model = st.selectbox("Ollama Model", models)
                    os.environ["OLLAMA_MODEL"] = ollama_model
                    st.success(f"✓ Ollama running · {len(models)} model(s)")
                else:
                    st.warning("Ollama running but no models pulled.\n`ollama pull llama3.2`")
            else:
                st.error("✗ Ollama offline. Run: `ollama serve`")
        except Exception:
            ollama_model = st.text_input("Ollama Model", value="llama3.2")
            os.environ["OLLAMA_MODEL"] = ollama_model

    os.environ["LLM_BACKEND"] = selected_backend
    from backend.llm.adapter import get_adapter, reset_adapter
    reset_adapter()

    # show active status
    try:
        _a = get_adapter()
        st.caption(f"Active: {_a.status_line()}")
    except Exception as e:
        st.caption(f"Config: {e}")

    st.divider()

    # ── Company / period ──────────────────────────────────────────────────────
    st.markdown("### Report Config")
    from data.sample_companies import COMPANIES
    co_options = {f"{v['_meta']['name']} ({v['_meta']['sector']})": k for k, v in COMPANIES.items()}
    co_options["Upload Excel"] = "_upload"
    co_options["Custom (manual entry)"] = "_custom"
    selected_co_label = st.selectbox("Company Dataset", list(co_options.keys()), index=0)
    selected_co_key   = co_options[selected_co_label]

    if selected_co_key == "_upload":
        st.download_button(
            "⬇ Download Template (.xlsx)",
            data=_make_template(),
            file_name="cfo_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        uploaded_file = st.file_uploader("Upload completed template", type=["xlsx", "xls"])
        company = st.text_input("Company Name", value="My Company Inc.")
        period  = st.text_input("Period", value="Q1 2026")
        if uploaded_file:
            try:
                active_data, n_fields = _parse_upload(uploaded_file.read())
                st.success(f"✓ {n_fields} fields loaded from Excel")
            except Exception as e:
                st.error(f"Parse error: {e}")
                active_data = DEFAULT_DATA
        else:
            st.caption("Download the template, fill in your numbers, then upload it here.")
            active_data = DEFAULT_DATA
    elif selected_co_key == "_custom":
        company = st.text_input("Company Name", value="My Company Inc.")
        period  = st.text_input("Period", value="Q1 2026")
        active_data = DEFAULT_DATA
    else:
        co_data = COMPANIES[selected_co_key]
        company = co_data["_meta"]["name"]
        period  = co_data["_meta"]["period"]
        active_data = co_data
        st.caption(f"Sector: {co_data['_meta']['sector']} · {co_data['_meta']['note']}")

    st.divider()

    # ── Run deterministic pipeline ────────────────────────────────────────────
    _auto_save_memory = st.checkbox(
        "Auto-save to Institutional Memory", value=True,
        help="Persist KPIs, anomalies, and GAAP/IFRS results for long-term trend tracking"
    )
    if st.button("▶  Run Deterministic Pipeline", type="primary"):
        with st.spinner("Running Math → GAAP → IFRS → RAG …"):
            try:
                st.session_state["results"] = run_deterministic(active_data, company, period)
                st.session_state["analysis"] = None
                st.session_state["debate"]   = None
                # Auto-persist to institutional memory
                if _auto_save_memory:
                    try:
                        from backend.memory.engine import get_memory_engine
                        _mem = get_memory_engine()
                        _sector = active_data.get("_meta", {}).get("sector", "")
                        _snap_id = _mem.save_analysis(
                            company_name=company,
                            period=period,
                            results=st.session_state["results"],
                            sector=_sector,
                        )
                        st.session_state["memory_snap_id"] = _snap_id
                        st.success(f"Pipeline complete! Memory saved (snap {_snap_id[:8]}…)")
                    except Exception as _me:
                        st.success("Pipeline complete!")
                        st.caption(f"Memory save skipped: {_me}")
                else:
                    st.success("Pipeline complete!")
            except Exception as e:
                st.error(f"Pipeline error: {e}")

    # ── LLM buttons ───────────────────────────────────────────────────────────
    if "results" in st.session_state:
        if st.button("🧠  Run AI Analysis"):
            with st.spinner("Generating CFO analysis …"):
                try:
                    from backend.agents.analysis_agent import analysis_agent_node
                    r = st.session_state["results"]
                    state = {
                        "company_name": r["company"], "period": r["period"],
                        "task_description": f"{r['period']} board analysis",
                        "task_type": "full_report", "report_format": "board",
                        "validated_data": r["data"],
                        "kpi_metrics": r["kpis"], "variance_table": r["variance"],
                        "anomaly_flags": r["anomalies"], "forecast_outputs": r["forecast"],
                        "gaap_results": r["gaap"], "ifrs_results": r["ifrs"],
                        "rag_chunks": r["rag_chunks"], "errors": [], "audit_log": [],
                        "agent_statuses": {},
                    }
                    out = analysis_agent_node(state, backend=selected_backend)
                    st.session_state["analysis"] = out
                    if out.get("errors"):
                        st.warning(f"Warnings: {out['errors']}")
                    else:
                        st.success("Analysis complete!")
                except Exception as e:
                    st.error(f"Analysis error: {e}")

        if st.button("⚖  Run GAAP/IFRS Debate"):
            with st.spinner("Running 3-round debate (this takes ~60–90s) …"):
                try:
                    from backend.agents.debate_agent import debate_agent_node
                    r = st.session_state["results"]
                    state = {
                        "company_name": r["company"], "period": r["period"],
                        "validated_data": r["data"],
                        "kpi_metrics": r["kpis"],
                        "gaap_results": r["gaap"], "ifrs_results": r["ifrs"],
                        "errors": [], "audit_log": [], "agent_statuses": {},
                    }
                    out = debate_agent_node(state, backend=selected_backend)
                    st.session_state["debate"] = out
                    if out.get("errors"):
                        st.warning(f"Warnings: {out['errors']}")
                    else:
                        st.success("Debate complete!")
                except Exception as e:
                    st.error(f"Debate error: {e}")

    st.divider()
    st.markdown("### 📊 Dashboards")
    if "results" in st.session_state:
        if st.button("🖥  Generate All Dashboards", type="primary"):
            with st.spinner("Generating 4 dashboards …"):
                try:
                    import subprocess

                    from dashboards.html_generators import (
                        generate_cfo_dashboard,
                        generate_cost_dashboard,
                        generate_headcount_dashboard,
                        generate_inventory_dashboard,
                    )
                    r = st.session_state["results"]
                    paths = []
                    paths.append(generate_cfo_dashboard(
                        r["data"], r["kpis"], r["variance"], r["gaap"], r["ifrs"],
                        r["forecast"], r["runway"], r["anomalies"], r["rag_chunks"],
                        r["company"], r["period"],
                    ))
                    paths.append(generate_cost_dashboard(r["data"], r["kpis"], r["company"], r["period"]))
                    paths.append(generate_headcount_dashboard(r["data"], r["kpis"], r["company"], r["period"]))
                    paths.append(generate_inventory_dashboard(r["data"], r["kpis"], r["company"], r["period"]))
                    for p in paths:
                        subprocess.Popen(["cmd", "/c", "start", "", p], shell=False)
                    st.success("✓ All 4 dashboards generated and opened!")
                    st.session_state["dashboard_paths"] = paths
                except Exception as e:
                    st.error(f"Dashboard error: {e}")

        if "dashboard_paths" in st.session_state:
            st.caption("Last generated:")
            names = ["AI CFO", "Cost Analysis", "Headcount KPI", "Inventory"]
            for name, path in zip(names, st.session_state["dashboard_paths"]):
                if st.button(f"↗ Open {name}", key=f"open_{name}"):
                    import subprocess
                    subprocess.Popen(["cmd", "/c", "start", "", path], shell=False)
    else:
        st.caption("Run the pipeline first to enable dashboards.")

    st.divider()
    st.caption("Layer 1: Pandas/NumPy/Sklearn · ZERO LLM\nLayer 2: Pydantic v2 schemas\nLayer 3: FastAPI + pgvector\nLayer 4: RAG retrieved context")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN AREA
# ══════════════════════════════════════════════════════════════════════════════
if "results" not in st.session_state:
    st.markdown("## Welcome to the AI CFO Multi-Agent System")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info("**Step 1** — Configure LLM backend in the sidebar (Anthropic or Ollama)")
    with col2:
        st.info("**Step 2** — Click **Run Deterministic Pipeline** to compute all KPIs, GAAP, IFRS, and RAG")
    with col3:
        st.info("**Step 3** — Click **Run AI Analysis** or **Run GAAP/IFRS Debate** for LLM interpretation")
    st.stop()

r = st.session_state["results"]
kpis = r["kpis"]
data = r["data"]

# ── header ───────────────────────────────────────────────────────────────────
gc = sum(1 for v in r["gaap"].values() if v.get("status") == "COMPLIANT")
ic = sum(1 for v in r["ifrs"].values() if v.get("status") == "COMPLIANT")

st.markdown(f"## {r['company']} — {r['period']}")
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Revenue",      fmt(data["revenue"]))
c2.metric("EBITDA Margin", f"{kpis['ebitda_margin_pct']:.1f}%")
c3.metric("Gross Margin",  f"{kpis['gross_margin_pct']:.1f}%")
c4.metric("GAAP",         f"{gc}/12 ✓")
c5.metric("IFRS",         f"{ic}/12 ✓")
st.divider()

# ── tabs ──────────────────────────────────────────────────────────────────────
tabs = st.tabs([
    "📈 Overview & KPIs",
    "📊 Variance Analysis",
    "⚖ Compliance",
    "🔮 Forecast & Segments",
    "↔ GAAP/IFRS Recon",
    "🧠 AI Analysis",
    "💬 Debate",
    "🔍 RAG & Audit",
    "💰 CapEx Analysis",
    "👤 HITL Approval",
    "🏦 Institutional Memory",
])


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 · OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
with tabs[0]:
    st.subheader("Income Statement KPIs")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Revenue",      fmt(data["revenue"]),    f"vs budget {fmt(data['budget']['revenue'])}")
    c2.metric("Gross Profit", fmt(data["gross_profit"]),f"GM {kpis['gross_margin_pct']:.1f}%")
    c3.metric("EBITDA",       fmt(data["ebitda"]),      f"Margin {kpis['ebitda_margin_pct']:.1f}%")
    c4.metric("Net Income",   fmt(data["net_income"]),  f"NM {kpis['net_margin_pct']:.1f}%")

    st.subheader("Balance Sheet & Liquidity")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Cash",           fmt(data["cash"]))
    c2.metric("Current Ratio",  f"{kpis['current_ratio']:.2f}x",
              "✓ Healthy" if kpis["current_ratio"] >= 2 else "⚠ Monitor")
    c3.metric("Net Debt",       fmt(kpis["net_debt"]))
    c4.metric("Working Capital",fmt(kpis["working_capital"]))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("ROE",             f"{kpis['roe_pct']:.1f}%")
    c2.metric("ROA",             f"{kpis['roa_pct']:.1f}%")
    c3.metric("D/E Ratio",       f"{kpis['debt_to_equity']:.2f}x")
    c4.metric("Diluted EPS",     f"${kpis['diluted_eps']:.2f}")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("DSO",     f"{kpis['dso_days']:.0f} days")
    c2.metric("CCC",     f"{kpis['ccc_days']:.0f} days")
    c3.metric("Int. Coverage", f"{kpis['interest_coverage']:.1f}x")
    c4.metric("Eff. Tax Rate",  f"{kpis['effective_tax_rate']:.1f}%")

    st.subheader("SaaS Metrics")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("ARR",    fmt(data["arr"]))
    c2.metric("NRR",    f"{data['nrr_pct']}%")
    c3.metric("Churn",  f"{data['churn_rate_pct']}%")
    c4.metric("Headcount", str(data["headcount"]))

    # Anomalies
    if r["anomalies"]:
        st.subheader("Anomaly Flags")
        for flag in r["anomalies"]:
            if "CRITICAL" in flag:
                st.error(flag)
            else:
                st.warning(flag)
    else:
        st.success("✓ No statistical anomalies detected — all indicators within normal bounds")

    # Cash runway
    runway = r["runway"]
    st.subheader("Cash Runway (ASC 205-40)")
    rm = runway["runway_months"]
    label = "∞ Profitable" if rm > 900 else f"{rm} months"
    if runway["status"] == "ADEQUATE":
        st.success(f"✓ {label} — ASC 205-40 not applicable (> 12 months)")
    elif runway["status"] == "WARNING":
        st.warning(f"⚠ {label} — Approaching 12-month going concern threshold")
    else:
        st.error(f"✗ {label} — ASC 205-40 going concern disclosure required")


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 · VARIANCE
# ══════════════════════════════════════════════════════════════════════════════
with tabs[1]:
    st.subheader("Budget vs Actuals — SAB 99 Materiality (≥ 5%)")

    rows = []
    labels = {"revenue": "Revenue", "cogs": "Cost of Revenue",
              "gross_profit": "Gross Profit", "ebitda": "EBITDA",
              "rd_expense": "R&D Expense", "sg_a": "SG&A Expense"}
    for key, label in labels.items():
        item = r["variance"]["line_items"].get(key, {})
        if not item:
            continue
        rows.append({
            "Line Item":    label,
            "Actual":       fmt(item["actual"]),
            "Budget":       fmt(item["budget"]),
            "Variance $":   fmt(item["variance_abs"]),
            "Variance %":   f"{item['variance_pct']:+.1f}%",
            "Favorable":    "✓" if item["favorable"] else "✗",
            "Material (SAB 99)": "⚠ Yes" if item["material"] else "No",
        })

    t = r["variance"]["totals"]
    rows.append({
        "Line Item": "TOTAL",
        "Actual":    fmt(t["actual"]),
        "Budget":    fmt(t["budget"]),
        "Variance $": fmt(t["variance_abs"]),
        "Variance %": f"{t['variance_pct']:+.1f}%",
        "Favorable":  "✓" if t["favorable"] else "✗",
        "Material (SAB 99)": "—",
    })

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Variance bar chart
    st.subheader("Variance Bridge")
    chart_data = {k: v["variance_abs"]
                  for k, v in r["variance"]["line_items"].items()
                  if k in labels}
    chart_df = pd.DataFrame.from_dict(
        {"Line Item": list(labels[k] for k in chart_data),
         "Variance": list(chart_data.values())}
    )
    st.bar_chart(chart_df.set_index("Line Item"))

    # ── SAB 99 AI Explanation ────────────────────────────────────────────────
    st.divider()
    st.subheader("SAB 99 AI Interpretation")
    st.caption("SEC Staff Accounting Bulletin No. 99 — Materiality threshold ≥ 5%")

    if st.button("🧠 Ask AI to Explain SAB 99 Variances", key="sab99_ai_btn"):
        with st.spinner("Generating SAB 99 materiality analysis …"):
            try:
                from backend.llm.adapter import get_adapter as _get_adapter
                _adapter = _get_adapter()
                _var_lines = []
                for _key, _label in labels.items():
                    _item = r["variance"]["line_items"].get(_key, {})
                    if _item:
                        _mat = "MATERIAL (≥5%)" if _item.get("material") else "not material"
                        _fav = "Favorable" if _item.get("favorable") else "Unfavorable"
                        _var_lines.append(
                            f"- {_label}: Actual {fmt(_item['actual'])}, Budget {fmt(_item['budget'])}, "
                            f"Variance {_item['variance_pct']:+.1f}% — {_mat} · {_fav}"
                        )
                _var_summary = "\n".join(_var_lines)
                _sab99_prompt = f"""You are a Lead Audit CPA specializing in SEC materiality standards.

Company: {r['company']} | Period: {r['period']}

**SAB 99 (SEC Staff Accounting Bulletin No. 99)** establishes that a 5% quantitative variance is a useful starting point but NOT a safe harbor. Qualitative factors can make smaller variances material.

Budget vs Actuals (current period):
{_var_summary}

Please provide:
1. **What is SAB 99?** — plain-English explanation for the CFO (2-3 sentences)
2. **Per-Category Analysis** — for each line item, explain:
   - Whether the variance is material under SAB 99 and why
   - Qualitative factors that may elevate or reduce materiality
   - Required SEC/board disclosure language if material
3. **Priority Actions** — ranked list of disclosures or footnotes required
4. **Overall Materiality Assessment** — is the aggregate variance pattern a concern?

Be specific with dollar amounts. Cite ASC standards and SAB 99 where relevant."""
                _sab99_response = _adapter.complete(
                    "You are a Lead Audit CPA specializing in SEC materiality standards under SAB 99.",
                    _sab99_prompt,
                    max_tokens=1800,
                )
                st.session_state["sab99_ai_response"] = _sab99_response
                st.success("SAB 99 analysis complete!")
            except Exception as _e:
                st.error(f"AI error: {_e}")

    if st.session_state.get("sab99_ai_response"):
        st.markdown(st.session_state["sab99_ai_response"])


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 3 · COMPLIANCE
# ══════════════════════════════════════════════════════════════════════════════
with tabs[2]:
    col_g, col_i = st.columns(2)

    gaap_labels = {
        "asc205": "ASC 205-40 Going Concern", "asc230": "ASC 230 Cash Flows",
        "asc260": "ASC 260 EPS",              "asc280": "ASC 280 Segments",
        "asc310": "ASC 310/326 CECL",         "asc350": "ASC 350 Goodwill",
        "asc450": "ASC 450 Contingencies",    "asc606": "ASC 606 Revenue",
        "asc740": "ASC 740 Income Taxes",     "asc820": "ASC 820 Fair Value",
        "asc842": "ASC 842 Leases",           "sab99":  "SAB 99 Materiality",
    }
    ifrs_labels = {
        "ias1":  "IAS 1 Presentation",  "ias2":  "IAS 2 Inventories",
        "ias7":  "IAS 7 Cash Flows",    "ias12": "IAS 12 Income Taxes",
        "ias16": "IAS 16 PPE",          "ias33": "IAS 33 EPS",
        "ias36": "IAS 36 Impairment",   "ias37": "IAS 37 Provisions",
        "ias38": "IAS 38 Intangibles",  "ifrs9": "IFRS 9 Credit Losses",
        "ifrs15":"IFRS 15 Revenue",     "ifrs16":"IFRS 16 Leases",
    }

    with col_g:
        gc = sum(1 for v in r["gaap"].values() if v.get("status") == "COMPLIANT")
        st.metric("GAAP Compliant", f"{gc}/12")
        gaap_rows = []
        for std, label in gaap_labels.items():
            res = r["gaap"].get(std, {})
            gaap_rows.append({
                "Standard": label,
                "Status":   res.get("status", "—"),
                "Finding":  res.get("finding", "—")[:90],
            })
        gdf = pd.DataFrame(gaap_rows)

        def color_status(val):
            c = {"COMPLIANT": "color: #10B981", "DISCLOSURE_REQUIRED": "color: #FBBF24",
                 "NON_COMPLIANT": "color: #F87171"}
            return c.get(val, "")

        st.dataframe(
            gdf.style.applymap(color_status, subset=["Status"]),
            use_container_width=True, hide_index=True,
        )

    with col_i:
        ic = sum(1 for v in r["ifrs"].values() if v.get("status") == "COMPLIANT")
        st.metric("IFRS Compliant", f"{ic}/12")
        ifrs_rows = []
        for std, label in ifrs_labels.items():
            res = r["ifrs"].get(std, {})
            ifrs_rows.append({
                "Standard": label,
                "Status":   res.get("status", "—"),
                "Finding":  res.get("finding", "—")[:90],
            })
        idf = pd.DataFrame(ifrs_rows)
        st.dataframe(
            idf.style.applymap(color_status, subset=["Status"]),
            use_container_width=True, hide_index=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 4 · FORECAST & SEGMENTS
# ══════════════════════════════════════════════════════════════════════════════
with tabs[3]:
    st.subheader("Revenue Forecast — LR 40% + Holt-Winters 60% Ensemble")

    hist   = data["historical_revenue"]
    fcast  = r["forecast"].get("forecast", [])
    q_hist = ["Q1'24","Q2'24","Q3'24","Q4'24","Q1'25","Q2'25","Q3'25","Q4'25","Q1'26"]
    q_fcast= ["Q2'26","Q3'26","Q4'26","Q1'27","Q2'27","Q3'27","Q4'27","Q1'28"]

    import altair as alt

    _hist_df = pd.DataFrame({
        "Quarter":        q_hist[:len(hist)],
        "Revenue ($M)":   [v / 1_000_000 for v in hist],
        "Type":           "Historical",
    })
    _fcast_vals = fcast[:len(q_fcast)] if fcast else []
    _fcast_df = pd.DataFrame({
        "Quarter":        q_fcast[:len(_fcast_vals)],
        "Revenue ($M)":   [v / 1_000_000 for v in _fcast_vals],
        "Type":           "Forecast",
    }) if _fcast_vals else pd.DataFrame(columns=["Quarter", "Revenue ($M)", "Type"])

    _color_scale = alt.Scale(domain=["Historical", "Forecast"], range=["#3B82F6", "#F59E0B"])
    _bars = (
        alt.Chart(_hist_df)
        .mark_bar(opacity=0.85, cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("Quarter:N", sort=None, axis=alt.Axis(labelAngle=-35, title=None)),
            y=alt.Y("Revenue ($M):Q", title="Revenue ($M)"),
            color=alt.Color("Type:N", scale=_color_scale, legend=alt.Legend(title=None)),
            tooltip=[alt.Tooltip("Quarter:N"), alt.Tooltip("Revenue ($M):Q", format="$.2f")],
        )
    )
    if not _fcast_df.empty:
        _fcast_line = (
            alt.Chart(_fcast_df)
            .mark_line(strokeDash=[6, 3], strokeWidth=2.5,
                       point=alt.OverlayMarkDef(filled=True, size=70))
            .encode(
                x=alt.X("Quarter:N", sort=None),
                y=alt.Y("Revenue ($M):Q"),
                color=alt.Color("Type:N", scale=_color_scale),
                tooltip=[alt.Tooltip("Quarter:N"), alt.Tooltip("Revenue ($M):Q", format="$.2f")],
            )
        )
        _rev_chart = (_bars + _fcast_line).properties(
            height=360, title="Revenue: Historical (bars) vs Forecast (dashed line)"
        )
    else:
        _rev_chart = _bars.properties(height=360, title="Historical Revenue")

    st.altair_chart(_rev_chart, use_container_width=True)

    if r["forecast"].get("r2") is not None:
        st.caption(
            f"R² = {r['forecast']['r2']:.4f} · "
            f"Method: {r['forecast'].get('method','')} · "
            f"Next Q: {fmt(fcast[0]) if fcast else 'N/A'}"
        )

    st.divider()
    st.subheader("Segment Analysis (ASC 280 — 10% threshold)")

    segs = data["segments"]
    seg_rows = []
    total_rev = sum(s["revenue"] for s in segs)
    for s in segs:
        gm = round(s["gross_profit"] / s["revenue"] * 100, 1) if s["revenue"] else 0
        seg_rows.append({
            "Segment":      s["name"],
            "Revenue":      fmt(s["revenue"]),
            "% of Total":   f"{s['revenue']/total_rev*100:.1f}%",
            "Gross Profit": fmt(s["gross_profit"]),
            "Gross Margin": f"{gm:.1f}%",
            "Assets":       fmt(s["assets"]),
            "Reportable?":  "✓ Yes" if s["revenue"]/total_rev >= 0.10 else "No",
        })
    st.dataframe(pd.DataFrame(seg_rows), use_container_width=True, hide_index=True)

    # Segment revenue bar chart
    _seg_chart_df = pd.DataFrame({
        "Segment": [s["name"] for s in segs],
        "Revenue ($M)": [s["revenue"] / 1_000_000 for s in segs],
        "Gross Margin %": [
            round(s["gross_profit"] / s["revenue"] * 100, 1) if s["revenue"] else 0
            for s in segs
        ],
    })
    _seg_bars = (
        alt.Chart(_seg_chart_df)
        .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
        .encode(
            x=alt.X("Segment:N", sort="-y", axis=alt.Axis(labelAngle=0)),
            y=alt.Y("Revenue ($M):Q", title="Revenue ($M)"),
            color=alt.Color("Segment:N", legend=None),
            tooltip=["Segment", "Revenue ($M)", "Gross Margin %"],
        )
        .properties(height=260, title="Segment Revenue Distribution")
    )
    st.altair_chart(_seg_bars, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 5 · GAAP↔IFRS RECONCILIATION
# ══════════════════════════════════════════════════════════════════════════════
with tabs[4]:
    st.subheader("GAAP-to-IFRS Key Reconciling Items")

    op_lease = data["operating_lease_expense"]
    rd       = data["rd_expense"]
    dev_pct  = data.get("rd_dev_capitalizable_pct", 0.35)

    recon_rows = [
        {
            "Topic":           "IFRS 16 vs ASC 842 — Leases",
            "GAAP Treatment":  "Dual model — operating leases in SG&A (ASC 842)",
            "IFRS Treatment":  "Single model — all leases capitalized (IFRS 16)",
            "P&L / BS Impact": fmt(op_lease),
            "Effect":          "EBITDA higher under IFRS",
        },
        {
            "Topic":           "IAS 38 vs ASC 730 — R&D",
            "GAAP Treatment":  "All R&D expensed immediately (ASC 730)",
            "IFRS Treatment":  f"~{dev_pct*100:.0f}% dev costs capitalizable (IAS 38)",
            "P&L / BS Impact": fmt(rd * dev_pct),
            "Effect":          "Net income higher under IFRS",
        },
        {
            "Topic":           "IAS 37 vs ASC 450 — Provisions",
            "GAAP Treatment":  "~75% probable threshold (ASC 450)",
            "IFRS Treatment":  ">50% probable threshold (IAS 37)",
            "P&L / BS Impact": "Qualitative",
            "Effect":          "Earlier recognition under IFRS",
        },
        {
            "Topic":           "IAS 36 vs ASC 350 — Goodwill",
            "GAAP Treatment":  "No impairment reversal (ASC 350)",
            "IFRS Treatment":  "Reversal permitted at CGU level (IAS 36)",
            "P&L / BS Impact": fmt(data["goodwill"]),
            "Effect":          "IFRS allows upside recovery",
        },
        {
            "Topic":           "IAS 2 vs ASC 330 — Inventory",
            "GAAP Treatment":  "LIFO / FIFO / Weighted avg (ASC 330)",
            "IFRS Treatment":  "LIFO STRICTLY PROHIBITED (IAS 2)",
            "P&L / BS Impact": "N/A — company uses FIFO",
            "Effect":          "Compliant under both",
        },
        {
            "Topic":           "IAS 7 vs ASC 230 — Cash Flows",
            "GAAP Treatment":  "Interest paid = Operating (mandatory, ASC 230)",
            "IFRS Treatment":  "Interest paid = Operating OR Financing (policy choice, IAS 7)",
            "P&L / BS Impact": fmt(data["interest_expense"]),
            "Effect":          "IFRS gives flexibility",
        },
    ]

    col1, col2 = st.columns(2)
    with col1:
        st.metric("IFRS EBITDA Uplift",
                  fmt(data["ebitda"] + op_lease),
                  f"+{fmt(op_lease)} vs GAAP EBITDA {fmt(data['ebitda'])}")
    with col2:
        st.metric("IAS 38 R&D Capitalisation Benefit",
                  fmt(rd * dev_pct),
                  f"{dev_pct*100:.0f}% of {fmt(rd)} R&D")

    st.dataframe(pd.DataFrame(recon_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 6 · AI ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with tabs[5]:
    analysis = st.session_state.get("analysis")

    if not analysis:
        st.info("Click **🧠 Run AI Analysis** in the sidebar to generate CFO interpretation.")
        st.caption("Works with Anthropic (API key) or Ollama (local model, no key needed).")
    else:
        s = analysis.get("structured_output") or analysis

        # Executive Summary
        st.subheader("Executive Summary")
        summary = s.get("analysis_narrative") or s.get("executive_summary", "")
        st.markdown(summary)

        col1, col2 = st.columns(2)
        with col1:
            conf = s.get("ai_confidence_score") or s.get("confidence_score", 0)
            st.metric("Confidence Score", f"{conf:.0%}")

        # Key Variance Drivers
        drivers = s.get("key_variance_drivers", [])
        if drivers:
            st.subheader("Key Variance Drivers")
            for d in drivers:
                st.markdown(f"- {d}")

        col1, col2 = st.columns(2)
        with col1:
            risks = s.get("identified_risks", [])
            if risks:
                st.subheader("Identified Risks")
                for risk in risks:
                    st.warning(risk)

        with col2:
            opps = s.get("opportunities", [])
            if opps:
                st.subheader("Opportunities")
                for opp in opps:
                    st.success(opp)

        # Action Items
        actions = s.get("action_items", [])
        if actions:
            st.subheader("Action Items (with Owners)")
            for i, a in enumerate(actions, 1):
                st.markdown(f"**{i}.** {a}")

        # Citations
        with st.expander("Citations & Sources"):
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown("**RAG Sources**")
                for c in s.get("rag_sources_cited", []):
                    st.caption(c)
            with col2:
                st.markdown("**GAAP Citations**")
                for c in s.get("gaap_citations", []):
                    st.caption(c)
            with col3:
                st.markdown("**IFRS Citations**")
                for c in s.get("ifrs_citations", []):
                    st.caption(c)

        # Errors if any
        if analysis.get("errors"):
            with st.expander("Warnings"):
                for e in analysis["errors"]:
                    st.caption(e)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 7 · DEBATE
# ══════════════════════════════════════════════════════════════════════════════
with tabs[6]:
    debate = st.session_state.get("debate")

    if not debate:
        st.info("Click **⚖ Run GAAP/IFRS Debate** in the sidebar to start the 3-round debate.")
        st.caption("Round 1: IFRS Advocate → Round 2: GAAP Advocate → Round 3: Independent Arbiter")
    else:
        # Round 1
        st.markdown(
            '<div class="round-header" style="border-color:#A78BFA">'
            '<strong style="color:#A78BFA">ROUND 1 — IFRS ADVOCATE</strong><br>'
            '<small>IASB expert · 25yr international experience</small></div>',
            unsafe_allow_html=True,
        )
        ifrs_arg = debate.get("debate_ifrs_advocate", "")
        st.markdown(ifrs_arg)

        st.divider()

        # Round 2
        st.markdown(
            '<div class="round-header" style="border-color:#60A5FA">'
            '<strong style="color:#60A5FA">ROUND 2 — GAAP ADVOCATE</strong><br>'
            '<small>FASB/SEC expert · 25yr NYSE/NASDAQ experience</small></div>',
            unsafe_allow_html=True,
        )
        gaap_arg = debate.get("debate_gaap_advocate", "")
        st.markdown(gaap_arg)

        st.divider()

        # Round 3
        st.markdown(
            '<div class="round-header" style="border-color:#10B981">'
            '<strong style="color:#10B981">ROUND 3 — INDEPENDENT ARBITER · VERDICT</strong><br>'
            '<small>IASB + FASB advisory council · No bias · Binding recommendation</small></div>',
            unsafe_allow_html=True,
        )
        arbiter = debate.get("debate_arbiter", "")
        st.markdown(arbiter)

        if debate.get("errors"):
            with st.expander("Warnings"):
                for e in debate["errors"]:
                    st.caption(e)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 8 · RAG & AUDIT
# ══════════════════════════════════════════════════════════════════════════════
with tabs[7]:
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("RAG — Retrieved Knowledge Chunks")
        st.caption(f"Query: {r['rag_query'][:120]}…")
        for i, chunk in enumerate(r["rag_chunks"], 1):
            with st.expander(f"[{i}] {chunk['title']} — score: {chunk.get('score', 0):.3f}"):
                st.markdown(chunk["content"])

    with col2:
        st.subheader("Pipeline Status")
        pipeline_status = {
            "data_agent":     ("COMPLETE", "✓"),
            "math_engine":    ("COMPLETE", "✓"),
            "rag_agent":      ("COMPLETE", "✓"),
            "gaap_agent":     ("COMPLETE", "✓"),
            "ifrs_agent":     ("COMPLETE", "✓"),
            "analysis_agent": ("COMPLETE" if st.session_state.get("analysis") and
                               not st.session_state["analysis"].get("errors") else "PENDING", ""),
            "debate_agent":   ("COMPLETE" if st.session_state.get("debate") and
                               not st.session_state["debate"].get("errors") else "PENDING", ""),
            "reporting_agent":("INACTIVE", ""),
        }
        for agent, (status, _) in pipeline_status.items():
            if status == "COMPLETE":
                st.success(f"✓ {agent}")
            elif status == "INACTIVE":
                st.warning(f"⊘ {agent} — requires full backend pipeline (POST /tasks)")
            else:
                st.info(f"○ {agent}")

        st.divider()
        st.subheader("LLM Backend")
        try:
            _a = get_adapter()
            st.code(_a.status_line())
        except Exception as e:
            st.caption(str(e))

        st.divider()
        st.subheader("Data Quality")
        req = ["revenue","cogs","gross_profit","ebitda","net_income","total_assets","total_equity","cash"]
        score = sum(1 for f in req if data.get(f)) / len(req)
        st.metric("Completeness", f"{score:.0%}")
        st.progress(score)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 9 · CAPEX ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with tabs[8]:
    import altair as alt

    st.subheader("Capital Expenditure Dashboard")

    _capex   = data.get("capex", 0)
    _ocf     = data.get("cash_from_operations", 0)
    _fcf     = data.get("free_cash_flow", 0)
    _rev     = data.get("revenue", 1)
    _assets  = data.get("total_assets", 1)
    _ebitda  = data.get("ebitda", 0)
    _depr    = data.get("depreciation", 0)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Capital Expenditures",  fmt(_capex))
    c2.metric("Free Cash Flow",        fmt(_fcf),   f"OCF − CapEx")
    c3.metric("CapEx / Revenue",       f"{_capex / _rev * 100:.1f}%")
    c4.metric("FCF Margin",            f"{_fcf / _rev * 100:.1f}%")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Operating Cash Flow",   fmt(_ocf))
    c2.metric("CapEx / OCF",           f"{_capex / _ocf * 100:.1f}%" if _ocf else "N/A")
    c3.metric("FCF / Assets",          f"{_fcf / _assets * 100:.1f}%")
    c4.metric("Maintenance CapEx Est.", fmt(_depr),  "≈ D&A proxy")

    st.divider()
    st.subheader("Cash Flow Waterfall: OCF → CapEx → FCF")

    _waterfall_df = pd.DataFrame({
        "Category":    ["Operating Cash Flow", "Capital Expenditures", "Free Cash Flow"],
        "Amount ($M)": [_ocf / 1e6, _capex / 1e6, _fcf / 1e6],
        "Color":       ["positive", "negative", "positive" if _fcf >= 0 else "negative"],
    })
    _capex_bar = (
        alt.Chart(_waterfall_df)
        .mark_bar(cornerRadiusTopLeft=5, cornerRadiusTopRight=5)
        .encode(
            x=alt.X("Category:N", sort=None, axis=alt.Axis(labelAngle=0, title=None)),
            y=alt.Y("Amount ($M):Q", title="Amount ($M)"),
            color=alt.Color(
                "Color:N",
                scale=alt.Scale(domain=["positive", "negative"], range=["#10B981", "#F87171"]),
                legend=None,
            ),
            tooltip=[alt.Tooltip("Category:N"), alt.Tooltip("Amount ($M):Q", format="$.2f")],
        )
        .properties(height=300)
    )
    st.altair_chart(_capex_bar, use_container_width=True)

    st.divider()
    st.subheader("CapEx Classification (ASC 360 / IAS 16 / ASC 350-40)")

    _capex_class = pd.DataFrame([
        {"Category": "Property, Plant & Equipment",  "Est. %": "45%",
         "Est. Amount": fmt(_capex * 0.45), "Standard": "ASC 360 / IAS 16",
         "Treatment": "Capitalize & depreciate over useful life"},
        {"Category": "Technology & Internal-Use SW",  "Est. %": "35%",
         "Est. Amount": fmt(_capex * 0.35), "Standard": "ASC 350-40 / IAS 38",
         "Treatment": "Capitalize dev phase; expense prelim & post-impl"},
        {"Category": "Leasehold Improvements",        "Est. %": "20%",
         "Est. Amount": fmt(_capex * 0.20), "Standard": "ASC 842 / IFRS 16",
         "Treatment": "Capitalize; amortize over shorter of lease or useful life"},
    ])
    st.dataframe(_capex_class, use_container_width=True, hide_index=True)

    # Growth vs Maintenance CapEx breakdown
    st.divider()
    _growth_capex = max(0, _capex - _depr)
    _maint_capex  = min(_capex, _depr)
    c1, c2 = st.columns(2)
    c1.metric("Maintenance CapEx (est.)", fmt(_maint_capex),
              "≈ D&A — keeps existing assets running")
    c2.metric("Growth CapEx (est.)",      fmt(_growth_capex),
              "CapEx above D&A — expands productive capacity")

    st.divider()
    st.subheader("AI CapEx & FCF Analysis")
    st.caption("AI interprets CapEx efficiency, FCF sustainability, and accounting treatment under ASC 360 / IAS 16.")

    if st.button("🧠 Run AI CapEx Analysis", key="capex_ai_btn"):
        with st.spinner("Generating CapEx analysis …"):
            try:
                from backend.llm.adapter import get_adapter as _get_adapter
                _adapter = _get_adapter()
                _gr = r["forecast"].get("growth_rate")
                _capex_prompt = f"""You are a Lead Audit CPA and CFO advisor.

Company: {r['company']} | Period: {r['period']}

Capital Expenditure Data:
- CapEx: {fmt(_capex)} ({_capex / _rev * 100:.1f}% of revenue)
- Operating Cash Flow: {fmt(_ocf)}
- Free Cash Flow: {fmt(_fcf)} (FCF margin {_fcf / _rev * 100:.1f}%)
- CapEx / OCF ratio: {_capex / _ocf * 100:.1f}% (healthy range: 20–40%)
- Maintenance CapEx estimate (≈ D&A): {fmt(_maint_capex)}
- Growth CapEx estimate: {fmt(_growth_capex)}
- Total Assets: {fmt(_assets)}
- EBITDA: {fmt(_ebitda)} (D&A: {fmt(_depr)})
- Revenue growth rate: {f"{_gr:.1%}" if _gr is not None else "see forecast tab"}

Please provide a structured CFO-level analysis:

**1. CapEx Efficiency Assessment**
- Is the CapEx/Revenue ratio appropriate for this business model?
- Is the CapEx/OCF ratio sustainable?
- Compare to SaaS/tech sector benchmarks

**2. Free Cash Flow Quality**
- FCF sustainability and quality assessment
- Working capital considerations
- FCF conversion rate (FCF / Net Income: {fmt(_fcf)} / {fmt(data.get('net_income', 1))})

**3. GAAP Accounting Treatment (ASC 360, ASC 350-40, ASC 842)**
- Capitalization vs expensing threshold considerations
- Impairment test obligations (ASC 360-10-35)
- Internal-use software capitalization (ASC 350-40) opportunity

**4. IFRS Accounting Treatment (IAS 16, IAS 38, IFRS 16)**
- Component accounting requirement (IAS 16.43)
- Development cost capitalization under IAS 38 (applies to {fmt(data.get('rd_expense', 0))} R&D)
- Difference vs GAAP treatment and P&L impact

**5. Strategic Recommendations**
- CapEx optimization actions for next quarter
- Depreciation schedule and EBITDA impact going forward
- Capital allocation priority ranking

Cite all relevant accounting standards with ASC/IFRS numbers."""

                _capex_response = _adapter.complete(
                    "You are a Lead Audit CPA and CFO advisor specializing in capital expenditure and free cash flow analysis.",
                    _capex_prompt,
                    max_tokens=2000,
                )
                st.session_state["capex_ai_response"] = _capex_response
                st.success("CapEx analysis complete!")
            except Exception as _e:
                st.error(f"AI error: {_e}")

    if st.session_state.get("capex_ai_response"):
        st.markdown(st.session_state["capex_ai_response"])


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 10 · HITL APPROVAL
# ══════════════════════════════════════════════════════════════════════════════
with tabs[9]:
    from datetime import datetime as _dt

    st.subheader("Human-in-the-Loop (HITL) Approval Gate")
    st.caption(
        "Deterministic triggers are checked against thresholds (variance ≥10%, gross margin <30%, "
        "GAAP/IFRS violations). If any trigger fires, CFO sign-off is required before report distribution."
    )

    # Compute triggers from current pipeline results
    from backend.agents.math_engine import FinancialCalculationEngine as _FCE
    _hitl_engine = _FCE()
    _hitl_triggers = _hitl_engine.check_approval_triggers(
        r["variance"], r["kpis"], r["anomalies"], r["gaap"], r["ifrs"]
    )
    _needs_approval = len(_hitl_triggers) > 0

    st.divider()

    if _needs_approval:
        st.error(
            f"⚠ **{len(_hitl_triggers)} Approval Trigger(s) Detected** — "
            "CFO review is required before report distribution."
        )

        st.markdown("#### Approval Triggers — Expand Each for Full Detail")
        for _t in _hitl_triggers:
            _sev    = _t.get("severity", "high")
            _reason = _t.get("reason", "trigger")
            _msg    = _t.get("message", "")
            _std    = _t.get("standard", "")
            _disc   = _t.get("requires_disclosure", False)

            _sev_icon  = "🔴 CRITICAL" if _sev == "critical" else "🟡 HIGH"
            _exp_label = f"{_sev_icon}  |  {_reason.replace('_', ' ').upper()}"

            with st.expander(_exp_label, expanded=True):
                if _sev == "critical":
                    st.error(f"{_msg}" + (" — **Disclosure Required**" if _disc else ""))
                else:
                    st.warning(f"{_msg}" + (" — Disclosure Required" if _disc else ""))
                if _std:
                    st.caption(f"Accounting Standard: {_std}")

                # ── VARIANCE EXCEEDS 10% ─────────────────────────────────────
                if _reason == "variance_exceeds_10pct":
                    st.markdown("---")
                    st.markdown("**Line-Item Variance Breakdown (Budget vs Actuals)**")
                    _var_items  = r["variance"].get("line_items", {})
                    _var_totals = r["variance"].get("totals", {})
                    _vlabels = {
                        "revenue": "Revenue", "cogs": "Cost of Revenue",
                        "gross_profit": "Gross Profit", "ebitda": "EBITDA",
                        "rd_expense": "R&D Expense", "sg_a": "SG&A Expense",
                    }
                    _vrows = []
                    for _vk, _vl in _vlabels.items():
                        _vi = _var_items.get(_vk, {})
                        if _vi:
                            _vrows.append({
                                "Line Item":       _vl,
                                "Actual":          fmt(_vi["actual"]),
                                "Budget":          fmt(_vi["budget"]),
                                "Variance $":      fmt(_vi["variance_abs"]),
                                "Variance %":      f"{_vi['variance_pct']:+.1f}%",
                                "Favorable":       "✓" if _vi["favorable"] else "✗",
                                "SAB 99 Material": "⚠ Yes" if _vi["material"] else "No",
                            })
                    _vrows.append({
                        "Line Item":       "TOTAL",
                        "Actual":          fmt(_var_totals.get("actual", 0)),
                        "Budget":          fmt(_var_totals.get("budget", 0)),
                        "Variance $":      fmt(_var_totals.get("variance_abs", 0)),
                        "Variance %":      f"{_var_totals.get('variance_pct', 0):+.1f}%",
                        "Favorable":       "✓" if _var_totals.get("favorable") else "✗",
                        "SAB 99 Material": "—",
                    })
                    if _vrows:
                        st.dataframe(pd.DataFrame(_vrows), use_container_width=True, hide_index=True)

                    _mat_items = r["variance"].get("material_items", [])
                    if _mat_items:
                        _mat_names = ", ".join(_vlabels.get(i, i) for i in _mat_items)
                        st.markdown(f"**SAB 99 Material Items (≥5%):** {_mat_names}")

                    _total_var_pct = abs(_var_totals.get("variance_pct", 0))
                    st.markdown("---")
                    st.markdown("**GAAP Disclosures Required:**")
                    st.markdown(f"""\
- **SAB 99 (SEC Materiality)**: Total variance of {_total_var_pct:.1f}% exceeds the 10% quantitative threshold. Both quantitative AND qualitative materiality assessment required in MD&A. Per SAB 99 §1, qualitative factors (management intent, trend impact, segment concentration) can elevate materiality below 5%.
- **SEC Reg S-K Item 303 (MD&A)**: Material changes in results of operations must be discussed and quantified. Provide period-over-period comparison with specific business drivers for each material line item.
- **ASC 250-10 (Accounting Changes & Estimates)**: Evaluate whether any variance is driven by a change in estimate or accounting policy requiring prospective or retrospective disclosure.
- **ASC 280-10-50 (Segment Reporting)**: If variance is concentrated in a specific reportable segment, disaggregated disclosure in segment footnotes is required with a management explanation.
- **Action**: Draft MD&A variance commentary for each material item; assign reviewer and set sign-off deadline before report distribution.\
""")

                # ── GROSS MARGIN BELOW 30% ────────────────────────────────────
                elif _reason == "gross_margin_below_30pct":
                    st.markdown("---")
                    _gm   = kpis.get("gross_margin_pct", 0)
                    _rev  = data.get("revenue", 1)
                    _cogs = data.get("cogs", 0)
                    st.markdown("**Gross Margin Analysis:**")
                    _c1g, _c2g, _c3g = st.columns(3)
                    _c1g.metric("Gross Margin",    f"{_gm:.1f}%",    "Threshold: 30%")
                    _c2g.metric("Revenue",         fmt(_rev))
                    _c3g.metric("COGS",            fmt(_cogs),       f"{_cogs / _rev * 100:.1f}% of Rev")

                    _segs = data.get("segments", [])
                    if _segs:
                        st.markdown("**Segment Gross Margin (ASC 280):**")
                        _sgrows = []
                        for _seg in _segs:
                            _sgm = round(_seg["gross_profit"] / _seg["revenue"] * 100, 1) if _seg["revenue"] else 0
                            _sgrows.append({
                                "Segment":         _seg["name"],
                                "Revenue":         fmt(_seg["revenue"]),
                                "Gross Profit":    fmt(_seg["gross_profit"]),
                                "Gross Margin %":  f"{_sgm:.1f}%",
                                "Below 30% Threshold": "⚠ Yes" if _sgm < 30 else "✓ No",
                            })
                        st.dataframe(pd.DataFrame(_sgrows), use_container_width=True, hide_index=True)

                    st.markdown("---")
                    st.markdown("**GAAP Disclosures Required:**")
                    st.markdown(f"""\
- **ASC 280-10-50 (Segment Reporting)**: Disaggregate gross margin by reportable segment. Disclose which segments are below 30% threshold and provide the cost structure explanation.
- **SAB 99 (Qualitative Materiality)**: Gross margin at {_gm:.1f}% is a qualitative materiality factor that elevates the significance of COGS variances regardless of their absolute dollar amount.
- **ASC 230 (Cash Flows — MD&A Liquidity)**: If low gross margin compresses operating cash flow, disclose adequacy of liquidity sources and management's remediation plan in the Liquidity section of MD&A.
- **ASC 205-40 (Going Concern)**: Evaluate whether a deteriorating gross margin trend constitutes a going concern indicator under the 12-month assessment window. Document the evaluation in board minutes.
- **Action**: Prepare segment margin footnote, update MD&A Liquidity section, confirm ASC 205-40 board evaluation is documented.\
""")

                # ── MULTIPLE ANOMALIES ────────────────────────────────────────
                elif _reason == "multiple_anomalies":
                    st.markdown("---")
                    st.markdown(f"**All {len(r['anomalies'])} Statistical Anomaly Flags Detected:**")
                    for _anom in r["anomalies"]:
                        if "CRITICAL" in _anom:
                            st.error(_anom)
                        else:
                            st.warning(_anom)

                    _has_going_concern = any("going concern" in a.lower() for a in r["anomalies"])
                    _has_liquidity     = any(
                        kw in a.lower() for a in r["anomalies"]
                        for kw in ("current ratio", "liquidity", "runway")
                    )
                    st.markdown("---")
                    st.markdown("**GAAP Disclosures Required:**")
                    st.markdown(f"""\
- **ASC 205-40 (Going Concern)**: {"Going concern indicators detected — management must evaluate whether conditions raise substantial doubt. Disclosure required in the financial statements and MD&A." if _has_going_concern else "Evaluate the full anomaly pattern for ASC 205-40 going concern conditions. Document the assessment in board minutes."}
- **ASC 275 (Risks and Uncertainties)**: Multiple statistical anomalies constitute risk concentrations requiring footnote disclosure of the nature of operations and key estimates subject to significant uncertainty.
- **SEC MD&A (Reg S-K Item 303)**: Discuss each material anomaly in the Results of Operations and Liquidity & Capital Resources sections. Quantify the trend and explain management's response plan.
- **{"ASC 230 / Liquidity: " + ("Current ratio anomaly — disclose available liquidity sources and adequacy assessment in MD&A Liquidity section." if _has_liquidity else "Ensure operating metric anomalies are addressed in the Liquidity section with specific remediation actions.")}
- **Action**: Map each anomaly flag to a specific MD&A paragraph, footnote, or board disclosure item. Assign owner and deadline.\
""")

                # ── GAAP / IFRS COMPLIANCE TRIGGERS ──────────────────────────
                elif _reason.startswith("gaap_") or _reason.startswith("ifrs_"):
                    _framework   = "GAAP" if _reason.startswith("gaap_") else "IFRS"
                    _status_type = "NON_COMPLIANT" if "NON_COMPLIANT" in _reason else "DISCLOSURE_REQUIRED"
                    _all_comp    = r["gaap"] if _framework == "GAAP" else r["ifrs"]

                    st.markdown("---")
                    st.markdown(f"**{_framework} Compliance Finding:**")
                    for _ckey, _cval in _all_comp.items():
                        if _cval.get("status") == _status_type and _std and (
                            _cval.get("standard") == _std
                            or _std in str(_cval.get("standard", ""))
                        ):
                            st.markdown(f"**Standard:** {_cval.get('standard', _std)}")
                            st.markdown(f"**Finding:** {_cval.get('finding', _msg)}")
                            _issues = _cval.get("issues", [])
                            if _issues:
                                st.markdown("**Issues identified:**")
                                for _iss in _issues:
                                    st.caption(f"• {_iss}")
                            break

                    st.markdown("---")
                    st.markdown("**GAAP Disclosures Required:**")
                    if _status_type == "NON_COMPLIANT":
                        st.markdown(f"""\
- **Immediate corrective action required** for {_std}: Non-compliance must be remediated before report issuance. Engage external auditors and legal counsel.
- **ASC 250-10 / IAS 8 (Restatement)**: If prior periods are affected, evaluate whether restatement is required and prepare restatement disclosure language.
- **Auditor communication**: Non-compliance must be disclosed to the external audit team immediately; risk of qualified or adverse opinion if unresolved.
- **SEC Form 8-K (Item 4.02)**: If the non-compliance relates to previously issued financial statements, evaluate whether Form 8-K disclosure is required under Item 4.02 (Non-Reliance on Previously Issued Financial Statements).
- **Action**: Escalate to external auditors within 24 hours; prepare remediation plan with specific ASC/IFRS remediation steps and sign-off timeline.\
""")
                    else:
                        st.markdown(f"""\
- **Footnote disclosure required** for {_std}: Prepare the specific footnote addressing this disclosure requirement. Assign a footnote number, primary drafter, and review deadline.
- **MD&A cross-reference**: Reference the footnote in the relevant MD&A section (Liquidity & Capital Resources, Critical Accounting Estimates, or Contractual Obligations as applicable to {_std}).
- **ASC 250 / IAS 8 (Consistency)**: Ensure disclosure language is consistent with prior period disclosures or document and disclose the reason for any policy change.
- **Auditor sign-off**: Share draft footnote language with external auditors for review before finalizing; confirm no additional disclosure is required.
- **Action**: Draft footnote, cross-reference in MD&A, confirm with auditors, and update disclosure checklist.\
""")

        st.divider()
        st.subheader("CFO Review Decision")

        _decision = st.session_state.get("hitl_decision", "pending")

        if _decision == "pending":
            st.info(
                "**Review checklist before approving:**\n"
                "1. All variance explanations are accurate and complete\n"
                "2. GAAP/IFRS disclosure requirements addressed in notes\n"
                "3. Management action plan reviewed and approved\n"
                "4. Report authorized for board/external distribution"
            )

            with st.form("hitl_approval_form"):
                _approver = st.text_input(
                    "Approver Name & Role",
                    placeholder="e.g. Jane Smith, CFO"
                )
                _notes = st.text_area(
                    "Review Notes (required — must address each trigger above)",
                    height=160,
                    placeholder=(
                        "1. Revenue variance of +11.7%: driven by enterprise upsell campaign — see deal desk memo\n"
                        "2. ASC 280 disclosure: added Note 14 with segment breakdown\n"
                        "3. Action owner: VP Finance (deadline: 2026-06-01)"
                    ),
                )
                _col1, _col2 = st.columns(2)
                with _col1:
                    _approve_btn = st.form_submit_button(
                        "✅ Approve — Authorize Distribution", type="primary"
                    )
                with _col2:
                    _reject_btn = st.form_submit_button("❌ Reject — Return for Revision")

                if _approve_btn:
                    if _approver and _notes:
                        st.session_state["hitl_decision"]  = "approved"
                        st.session_state["hitl_approver"]  = _approver
                        st.session_state["hitl_notes"]     = _notes
                        st.session_state["hitl_timestamp"] = _dt.utcnow().isoformat() + "Z"
                        try:
                            from backend.memory.engine import get_memory_engine as _gme
                            _gme().update_hitl(r["company"], r["period"], "approved", _approver, _notes)
                        except Exception:
                            pass
                        st.rerun()
                    else:
                        st.warning("Please fill in both Approver Name and Review Notes.")

                if _reject_btn:
                    if _approver and _notes:
                        st.session_state["hitl_decision"]  = "rejected"
                        st.session_state["hitl_approver"]  = _approver
                        st.session_state["hitl_notes"]     = _notes
                        st.session_state["hitl_timestamp"] = _dt.utcnow().isoformat() + "Z"
                        try:
                            from backend.memory.engine import get_memory_engine as _gme
                            _gme().update_hitl(r["company"], r["period"], "rejected", _approver, _notes)
                        except Exception:
                            pass
                        st.rerun()
                    else:
                        st.warning("Please fill in both Approver Name and Review Notes.")

        elif _decision == "approved":
            st.success(
                f"✅ **APPROVED** by **{st.session_state.get('hitl_approver', '')}** "
                f"at {st.session_state.get('hitl_timestamp', '')} UTC"
            )
            st.markdown(f"**Review Notes:** {st.session_state.get('hitl_notes', '')}")
            if st.button("🔄 Reset Decision", key="hitl_reset_btn"):
                for _k in ["hitl_decision", "hitl_approver", "hitl_notes", "hitl_timestamp"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        elif _decision == "rejected":
            st.error(
                f"❌ **REJECTED** by **{st.session_state.get('hitl_approver', '')}** "
                f"at {st.session_state.get('hitl_timestamp', '')} UTC — Report returned for revision."
            )
            st.markdown(f"**Review Notes:** {st.session_state.get('hitl_notes', '')}")
            if st.button("🔄 Reset Decision", key="hitl_reset_btn"):
                for _k in ["hitl_decision", "hitl_approver", "hitl_notes", "hitl_timestamp"]:
                    st.session_state.pop(_k, None)
                st.rerun()

    else:
        st.success(
            "✅ **No approval triggers detected** — Report may be distributed without CFO sign-off."
        )
        st.info(
            "Approval triggers fire automatically when:\n"
            "- Total variance vs budget **≥ 10%** (SAB 99 elevated materiality)\n"
            "- Gross margin falls **below 30%**\n"
            "- **3 or more** statistical anomalies detected\n"
            "- Any GAAP standard flagged **NON_COMPLIANT** or **DISCLOSURE_REQUIRED**\n"
            "- Any IFRS standard flagged **NON_COMPLIANT** or **DISCLOSURE_REQUIRED**"
        )

    # Audit log
    st.divider()
    with st.expander("HITL Audit Log (this session)"):
        _log_rows = []
        if st.session_state.get("hitl_decision"):
            _log_rows.append({
                "Timestamp":   st.session_state.get("hitl_timestamp", "—"),
                "Decision":    st.session_state.get("hitl_decision", "—").upper(),
                "Approver":    st.session_state.get("hitl_approver", "—"),
                "Triggers #":  len(_hitl_triggers),
                "Notes":       (st.session_state.get("hitl_notes", "—") or "")[:80],
            })
        if _log_rows:
            st.dataframe(pd.DataFrame(_log_rows), use_container_width=True, hide_index=True)
        else:
            st.caption("No approval actions recorded this session.")


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 11 · INSTITUTIONAL MEMORY
# ══════════════════════════════════════════════════════════════════════════════
with tabs[10]:
    import altair as alt

    st.subheader("Institutional Memory — Accumulated Company Knowledge")
    st.caption(
        "KPIs, compliance history, HITL decisions, and AI insights are persisted across "
        "every pipeline run — building an institutional record that grows richer over time."
    )

    try:
        from backend.memory.engine import get_memory_engine as _gme
        _mem_eng = _gme()

        # ── Company selector ─────────────────────────────────────────────────
        _all_cos = _mem_eng.get_all_companies()

        if not _all_cos:
            st.info(
                "No institutional memory yet.  Run the **Deterministic Pipeline** with "
                "**Auto-save to Institutional Memory** enabled to start building the record."
            )
        else:
            # Summary cards
            st.markdown(f"**{len(_all_cos)} companies in memory**")
            _co_cols = st.columns(min(len(_all_cos), 4))
            for _ci, _co in enumerate(_all_cos[:4]):
                with _co_cols[_ci]:
                    _n_analyses = _co["analysis_count"]
                    st.metric(
                        _co["company_name"][:22],
                        f"{_n_analyses} period{'s' if _n_analyses != 1 else ''}",
                        f"GM {_co['avg_gross_margin']:.1f}% avg",
                    )

            st.divider()

            # Company deep-dive
            _co_names = [c["company_name"] for c in _all_cos]
            _default_idx = _co_names.index(r["company"]) if r["company"] in _co_names else 0
            _selected_co = st.selectbox(
                "Select company to explore",
                _co_names,
                index=_default_idx,
                key="mem_co_select",
            )
            _co_record = next(c for c in _all_cos if c["company_name"] == _selected_co)
            _snapshots = _mem_eng.get_company_snapshots(_selected_co)

            # ── KPI Summary bar ──────────────────────────────────────────────
            _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns(5)
            _mc1.metric("Periods tracked",  _co_record["analysis_count"])
            _mc2.metric("Avg Gross Margin", f"{_co_record['avg_gross_margin']:.1f}%")
            _mc3.metric("Avg EBITDA Margin",f"{_co_record['avg_ebitda_margin']:.1f}%")
            _mc4.metric("Avg Net Margin",   f"{_co_record['avg_net_margin']:.1f}%")
            _mc5.metric("HITL Approval Rate", f"{_co_record['hitl_approval_rate']:.0f}%")

            # ── KPI Trend Charts ─────────────────────────────────────────────
            st.divider()
            st.subheader("KPI Trends Over Time")
            _trend_kpis = ["gross_margin_pct", "ebitda_margin_pct", "net_margin_pct",
                           "current_ratio", "debt_to_equity", "dso_days"]
            _trends = _mem_eng.get_kpi_trends(_selected_co, _trend_kpis)

            if any(_trends.values()):
                _trend_rows = []
                for _kname, _pts in _trends.items():
                    for _pt in _pts:
                        _trend_rows.append({
                            "Period":   _pt["period"],
                            "Value":    _pt["value"],
                            "KPI":      _kname.replace("_pct", " %").replace("_", " ").title(),
                            "Unit":     _pt["unit"],
                        })
                _trend_df = pd.DataFrame(_trend_rows)

                _kpi_labels = _trend_df["KPI"].unique().tolist()
                _selected_kpis = st.multiselect(
                    "KPIs to chart",
                    _kpi_labels,
                    default=_kpi_labels[:3],
                    key="mem_kpi_select",
                )
                if _selected_kpis:
                    _chart_df = _trend_df[_trend_df["KPI"].isin(_selected_kpis)]
                    _trend_chart = (
                        alt.Chart(_chart_df)
                        .mark_line(point=alt.OverlayMarkDef(filled=True, size=80))
                        .encode(
                            x=alt.X("Period:N", sort=None, axis=alt.Axis(labelAngle=-30, title=None)),
                            y=alt.Y("Value:Q", title="Value"),
                            color=alt.Color("KPI:N", legend=alt.Legend(title=None)),
                            tooltip=["Period", "KPI", alt.Tooltip("Value:Q", format=".2f")],
                        )
                        .properties(height=320, title="KPI Trend — All Tracked Periods")
                    )
                    st.altair_chart(_trend_chart, use_container_width=True)
            else:
                st.info("No KPI time-series data yet — run more pipeline periods.")

            # ── Period-over-Period Comparison Table ──────────────────────────
            st.divider()
            st.subheader("Period-over-Period Comparison")
            if len(_snapshots) >= 2:
                _pop_rows = []
                _pop_kpis = ["gross_margin_pct","ebitda_margin_pct","net_margin_pct",
                             "current_ratio","debt_to_equity","dso_days","ccc_days"]
                _pop_labels = {
                    "gross_margin_pct":  "Gross Margin %",
                    "ebitda_margin_pct": "EBITDA Margin %",
                    "net_margin_pct":    "Net Margin %",
                    "current_ratio":     "Current Ratio",
                    "debt_to_equity":    "Debt / Equity",
                    "dso_days":          "DSO (days)",
                    "ccc_days":          "CCC (days)",
                }
                _latest = _snapshots[-1]
                _prev   = _snapshots[-2]
                _lk = _latest["kpi_metrics"]
                _pk = _prev["kpi_metrics"]
                for _k, _label in _pop_labels.items():
                    _lv = _lk.get(_k)
                    _pv = _pk.get(_k)
                    if _lv is not None and _pv is not None:
                        _delta = round(_lv - _pv, 2)
                        _dpct  = round(_delta / abs(_pv) * 100, 1) if _pv else 0.0
                        _pop_rows.append({
                            "KPI":           _label,
                            f"{_prev['period']}": round(_pv, 2),
                            f"{_latest['period']}": round(_lv, 2),
                            "Delta":         f"{_delta:+.2f}",
                            "Delta %":       f"{_dpct:+.1f}%",
                            "Trend":         "↑" if _delta > 0 else ("↓" if _delta < 0 else "→"),
                        })
                if _pop_rows:
                    st.dataframe(pd.DataFrame(_pop_rows), use_container_width=True, hide_index=True)
            elif len(_snapshots) == 1:
                st.info("Only one period stored — run another period to see period-over-period comparison.")
            else:
                st.info("No snapshots yet.")

            # ── All Periods KPI History ──────────────────────────────────────
            st.divider()
            st.subheader("Full KPI History by Period")
            if _snapshots:
                _hist_rows = []
                for _s in _snapshots:
                    _k = _s["kpi_metrics"]
                    _hist_rows.append({
                        "Period":       _s["period"],
                        "Analyzed":     str(_s["analyzed_at"])[:10] if _s["analyzed_at"] else "—",
                        "GM %":         f"{_k.get('gross_margin_pct',0):.1f}%",
                        "EBITDA %":     f"{_k.get('ebitda_margin_pct',0):.1f}%",
                        "NM %":         f"{_k.get('net_margin_pct',0):.1f}%",
                        "Curr. Ratio":  f"{_k.get('current_ratio',0):.2f}x",
                        "DSO":          f"{_k.get('dso_days',0):.0f}d",
                        "Anomalies":    len(_s["anomaly_flags"]),
                        "HITL":         _s["hitl_decision"].upper(),
                        "Approver":     _s["hitl_approver"] or "—",
                    })
                st.dataframe(pd.DataFrame(_hist_rows), use_container_width=True, hide_index=True)

            # ── Peer Benchmarks ──────────────────────────────────────────────
            if len(_all_cos) > 1:
                st.divider()
                st.subheader("Peer Benchmarks")
                _peer = _mem_eng.get_peer_benchmarks(exclude_company=_selected_co)
                _pb1, _pb2, _pb3, _pb4 = st.columns(4)
                _pb1.metric("Peer Avg Gross Margin",
                            f"{_peer.get('avg_gross_margin',0):.1f}%",
                            f"{_co_record['avg_gross_margin'] - _peer.get('avg_gross_margin',0):+.1f}% vs you")
                _pb2.metric("Peer Avg EBITDA Margin",
                            f"{_peer.get('avg_ebitda_margin',0):.1f}%",
                            f"{_co_record['avg_ebitda_margin'] - _peer.get('avg_ebitda_margin',0):+.1f}% vs you")
                _pb3.metric("Peer Avg Net Margin",
                            f"{_peer.get('avg_net_margin',0):.1f}%",
                            f"{_co_record['avg_net_margin'] - _peer.get('avg_net_margin',0):+.1f}% vs you")
                _pb4.metric("Peer HITL Approval Rate",
                            f"{_peer.get('avg_approval_rate',0):.0f}%",
                            f"{_co_record['hitl_approval_rate'] - _peer.get('avg_approval_rate',0):+.0f}% vs you")

                # Peer comparison bar chart
                _peer_chart_data = []
                for _pco in _all_cos:
                    _peer_chart_data.append({
                        "Company":      _pco["company_name"][:20],
                        "Gross Margin %": _pco["avg_gross_margin"],
                        "EBITDA Margin %": _pco["avg_ebitda_margin"],
                        "Is Selected":  _pco["company_name"] == _selected_co,
                    })
                _pcd = pd.DataFrame(_peer_chart_data)
                _peer_bars = (
                    alt.Chart(_pcd)
                    .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                    .encode(
                        x=alt.X("Company:N", sort="-y", axis=alt.Axis(labelAngle=-30, title=None)),
                        y=alt.Y("Gross Margin %:Q", title="Avg Gross Margin %"),
                        color=alt.condition(
                            alt.datum["Is Selected"],
                            alt.value("#F59E0B"),
                            alt.value("#3B82F6"),
                        ),
                        tooltip=["Company", "Gross Margin %", "EBITDA Margin %"],
                    )
                    .properties(height=260, title="Peer Gross Margin Comparison (you = amber)")
                )
                st.altair_chart(_peer_bars, use_container_width=True)

            # ── Recurring Issues ─────────────────────────────────────────────
            st.divider()
            st.subheader("Recurring Issues & Compliance History")
            _rec_anom = _co_record.get("recurring_anomalies", [])
            _rec_comp = _co_record.get("compliance_issues", [])

            _ri1, _ri2 = st.columns(2)
            with _ri1:
                st.markdown("**Recurring Anomalies (≥ 2 periods)**")
                if _rec_anom:
                    for _a in _rec_anom:
                        if "CRITICAL" in _a:
                            st.error(_a)
                        else:
                            st.warning(_a)
                else:
                    st.success("No recurring anomalies.")

            with _ri2:
                st.markdown("**Recurring Compliance Issues (≥ 2 periods)**")
                if _rec_comp:
                    for _c in _rec_comp:
                        st.warning(f"⚠ {_c} — appeared in multiple periods")
                else:
                    st.success("No recurring compliance issues.")

            # ── Active Insights Timeline ─────────────────────────────────────
            st.divider()
            st.subheader("Active Insights Timeline")
            _insights = _mem_eng.get_insights(_selected_co, active_only=True)
            if _insights:
                _ins_df = pd.DataFrame([{
                    "Period":   i["period"],
                    "Type":     i["insight_type"].capitalize(),
                    "Severity": i["severity"].upper(),
                    "Source":   i["source"],
                    "Finding":  i["content"][:120],
                    "Logged":   str(i["created_at"])[:10] if i["created_at"] else "—",
                } for i in _insights])
                st.dataframe(_ins_df, use_container_width=True, hide_index=True)
            else:
                st.success("No active insights recorded.")

            # ── HITL Decision Log ────────────────────────────────────────────
            st.divider()
            st.subheader("HITL Decision Log (all periods)")
            _hitl_snaps = [s for s in _snapshots if s["hitl_decision"] != "pending"]
            if _hitl_snaps:
                _hlog = pd.DataFrame([{
                    "Period":    s["period"],
                    "Decision":  s["hitl_decision"].upper(),
                    "Approver":  s["hitl_approver"] or "—",
                    "Triggers":  len(s["approval_triggers"]),
                    "Notes":     (s["hitl_notes"] or "")[:80],
                    "Timestamp": str(s["hitl_timestamp"])[:16] if s["hitl_timestamp"] else "—",
                } for s in _hitl_snaps])
                st.dataframe(_hlog, use_container_width=True, hide_index=True)
            else:
                st.info("No HITL decisions recorded yet.")

            # ── AI Institutional Knowledge Synthesis ─────────────────────────
            st.divider()
            st.subheader("AI Institutional Knowledge Synthesis")
            _existing_summary = _co_record.get("institutional_summary", "")
            if _existing_summary:
                st.markdown(_existing_summary)
            else:
                st.info("No synthesis yet — click the button below to generate.")

            if st.button("🧠 Synthesize Institutional Knowledge", key="mem_synth_btn"):
                with st.spinner("Generating institutional knowledge summary …"):
                    try:
                        from backend.llm.adapter import get_adapter as _ga
                        _synth = _mem_eng.synthesize_knowledge(_selected_co, _ga())
                        st.markdown(_synth)
                        st.success("Institutional knowledge synthesized and saved!")
                    except Exception as _se:
                        st.error(f"Synthesis error: {_se}")

            # ── Export ───────────────────────────────────────────────────────
            st.divider()
            if _snapshots:
                _export_rows = []
                for _s in _snapshots:
                    _k = _s["kpi_metrics"]
                    _export_rows.append({
                        "company":       _selected_co,
                        "period":        _s["period"],
                        "analyzed_at":   str(_s["analyzed_at"])[:19] if _s["analyzed_at"] else "",
                        "gross_margin":  _k.get("gross_margin_pct", ""),
                        "ebitda_margin": _k.get("ebitda_margin_pct", ""),
                        "net_margin":    _k.get("net_margin_pct", ""),
                        "current_ratio": _k.get("current_ratio", ""),
                        "dso_days":      _k.get("dso_days", ""),
                        "anomaly_count": len(_s["anomaly_flags"]),
                        "hitl_decision": _s["hitl_decision"],
                        "hitl_approver": _s["hitl_approver"],
                    })
                _export_csv = pd.DataFrame(_export_rows).to_csv(index=False)
                st.download_button(
                    "⬇ Export Memory as CSV",
                    data=_export_csv,
                    file_name=f"institutional_memory_{_selected_co.replace(' ', '_')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

    except Exception as _mem_tab_err:
        st.error(f"Memory tab error: {_mem_tab_err}")
        import traceback
        st.code(traceback.format_exc())
