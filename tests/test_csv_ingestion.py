"""
Tests for CSV and Excel financial data ingestion.

Covers:
- Two-column CSV (field, value) format
- Wide CSV (headers + data row)
- Excel template format (Field / Value / Description columns)
- Alias normalisation
- Budget/actuals dict reconstruction
- Historical revenue list reconstruction
- Metadata injection (company_name, period, currency)
- ingest_bytes() for Streamlit upload path
- Error handling
"""
import io
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from backend.ingestion.csv_ingester import (
    ingest_csv,
    ingest_excel,
    ingest_file,
    ingest_bytes,
    _normalise_key,
    _coerce_field,
)


# ── Key normalisation ─────────────────────────────────────────────────────────

class TestNormaliseKey:
    def test_lowercase(self):
        assert _normalise_key("Revenue") == "revenue"

    def test_spaces_to_underscores(self):
        assert _normalise_key("Gross Profit") == "gross_profit"

    def test_hyphen_to_underscore(self):
        assert _normalise_key("pre-tax-income") == "pre_tax_income"

    def test_alias_rd(self):
        assert _normalise_key("R&D") == "rd_expense"

    def test_alias_sga(self):
        assert _normalise_key("SG&A") == "sg_a_expense"

    def test_alias_revenues(self):
        assert _normalise_key("revenues") == "revenue"

    def test_alias_cost_of_revenue(self):
        assert _normalise_key("Cost of Revenue") == "cogs"

    def test_alias_fcf(self):
        assert _normalise_key("FCF") == "free_cash_flow"

    def test_alias_employees(self):
        assert _normalise_key("employees") == "headcount"

    def test_strip_special_chars(self):
        key = _normalise_key("  Net Income!  ")
        assert key == "net_income"


# ── Value coercion ────────────────────────────────────────────────────────────

class TestCoerceField:
    def test_numeric_field_from_int(self):
        assert _coerce_field("revenue", 5_000_000) == 5_000_000.0

    def test_numeric_field_from_string_with_commas(self):
        assert _coerce_field("revenue", "5,000,000") == 5_000_000.0

    def test_numeric_field_from_dollar_string(self):
        assert _coerce_field("revenue", "$5,000,000") == 5_000_000.0

    def test_numeric_field_from_percent_string(self):
        assert _coerce_field("nrr_pct", "122%") == 122.0

    def test_none_returns_none(self):
        assert _coerce_field("revenue", None) is None

    def test_empty_string_returns_none(self):
        assert _coerce_field("revenue", "") is None

    def test_string_field_kept_as_string(self):
        assert _coerce_field("company_name", "Acme Corp") == "Acme Corp"

    def test_period_field_kept_as_string(self):
        assert _coerce_field("period", "Q1 2026") == "Q1 2026"

    def test_non_numeric_for_numeric_field_returns_none(self):
        assert _coerce_field("revenue", "not-a-number") is None


# ── CSV two-column format ─────────────────────────────────────────────────────

class TestCSVTwoColumn:
    def _make_csv(self, rows: list) -> io.BytesIO:
        text = "\n".join(f"{k},{v}" for k, v in rows)
        return io.BytesIO(text.encode())

    def test_basic_fields_parsed(self):
        data = ingest_csv(
            io.StringIO("company_name,Acme Corp\nperiod,Q1 2026\nrevenue,5000000\ncurrency,USD\n"),
        )
        assert data["company_name"] == "Acme Corp"
        assert data["period"] == "Q1 2026"
        assert data["revenue"] == 5_000_000.0

    def test_metadata_override(self):
        data = ingest_csv(
            io.StringIO("revenue,5000000\ncogs,2000000\n"),
            company_name="Override Co",
            period="Q2 2026",
            currency="EUR",
        )
        assert data["company_name"] == "Override Co"
        assert data["period"] == "Q2 2026"
        assert data["currency"] == "EUR"

    def test_currency_default_usd(self):
        data = ingest_csv(io.StringIO("revenue,1000000\n"))
        assert data["currency"] == "USD"

    def test_alias_rnd_parsed(self):
        data = ingest_csv(io.StringIO("R&D,2800000\n"))
        assert data.get("rd_expense") == 2_800_000.0

    def test_dollar_value_stripped(self):
        data = ingest_csv(io.StringIO("revenue,\"$5,000,000\"\n"))
        assert data["revenue"] == 5_000_000.0

    def test_budget_actuals_reconstructed(self):
        data = ingest_csv(io.StringIO(
            "actuals_revenue,15000000\nactuals_cogs,4500000\n"
            "budget_revenue,14000000\nbudget_cogs,4200000\n"
        ))
        assert "actuals" in data
        assert data["actuals"]["revenue"] == 15_000_000.0
        assert "budget" in data
        assert data["budget"]["revenue"] == 14_000_000.0
        # Raw actuals_* keys should be removed
        assert "actuals_revenue" not in data

    def test_historical_revenue_reconstructed(self):
        data = ingest_csv(io.StringIO(
            "hist_q1,1000000\nhist_q2,1100000\nhist_q3,1200000\n"
        ))
        assert "historical_revenue" in data
        assert data["historical_revenue"] == [1_000_000.0, 1_100_000.0, 1_200_000.0]
        assert "hist_q1" not in data

    def test_empty_csv_raises(self):
        with pytest.raises(ValueError, match="empty"):
            ingest_csv(io.StringIO(""))

    def test_from_bytes(self):
        data = ingest_csv(b"revenue,5000000\ncurrency,USD\n")
        assert data["revenue"] == 5_000_000.0

    def test_from_file_path(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("revenue,5000000\ncurrency,USD\ncompany_name,Test Co\nperiod,Q1 2026\n")
        data = ingest_csv(str(f))
        assert data["revenue"] == 5_000_000.0


# ── CSV wide format ───────────────────────────────────────────────────────────

class TestCSVWideFormat:
    def test_wide_format_parsed(self):
        csv_text = "revenue,cogs,net_income,currency,company_name,period\n15000000,4500000,2316000,USD,WideCo,Q1 2026\n"
        data = ingest_csv(io.StringIO(csv_text))
        assert data["revenue"] == 15_000_000.0
        assert data["net_income"] == 2_316_000.0
        assert data["company_name"] == "WideCo"

    def test_wide_format_header_alias(self):
        csv_text = "revenues,Cost of Revenue,FCF\n5000000,2000000,3000000\n"
        data = ingest_csv(io.StringIO(csv_text))
        assert data["revenue"] == 5_000_000.0
        assert data["cogs"] == 2_000_000.0
        assert data["free_cash_flow"] == 3_000_000.0


# ── Excel ingestion ───────────────────────────────────────────────────────────

class TestExcelIngestion:
    def _make_excel(self, rows: list) -> io.BytesIO:
        """Create a minimal Excel file with Field/Value columns."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Field", "Value", "Description"])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_excel_basic_fields(self):
        buf = self._make_excel([
            ["revenue", 15_000_000, "Net revenue"],
            ["cogs",     4_500_000, "COGS"],
            ["currency",     "USD", "Currency"],
            ["company_name", "ExcelCo", "Company"],
            ["period",    "Q1 2026", "Period"],
        ])
        data = ingest_excel(buf)
        assert data["revenue"] == 15_000_000.0
        assert data["cogs"] == 4_500_000.0
        assert data["company_name"] == "ExcelCo"

    def test_excel_section_headers_skipped(self):
        buf = self._make_excel([
            ["INCOME STATEMENT", None, None],
            ["revenue", 5_000_000, "Revenue"],
            ["BALANCE SHEET", None, None],
            ["total_assets", 50_000_000, "Assets"],
        ])
        data = ingest_excel(buf)
        assert "income_statement" not in data
        assert "balance_sheet" not in data
        assert data["revenue"] == 5_000_000.0

    def test_excel_metadata_override(self):
        buf = self._make_excel([
            ["revenue", 5_000_000, ""],
        ])
        data = ingest_excel(buf, company_name="Override", period="FY 2026", currency="EUR")
        assert data["company_name"] == "Override"
        assert data["period"] == "FY 2026"
        assert data["currency"] == "EUR"

    def test_excel_budget_actuals(self):
        buf = self._make_excel([
            ["actuals_revenue", 15_000_000, "Actual revenue"],
            ["budget_revenue", 14_000_000, "Budget revenue"],
        ])
        data = ingest_excel(buf)
        assert data["actuals"]["revenue"] == 15_000_000.0
        assert data["budget"]["revenue"] == 14_000_000.0

    def test_excel_historical_revenue(self):
        buf = self._make_excel([
            ["hist_q1", 10_000_000, "Q1 oldest"],
            ["hist_q2", 11_000_000, "Q2"],
            ["hist_q3", 12_000_000, "Q3 most recent"],
        ])
        data = ingest_excel(buf)
        assert data["historical_revenue"] == [10_000_000.0, 11_000_000.0, 12_000_000.0]

    def test_excel_from_file_path(self, tmp_path):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Field", "Value", "Desc"])
        ws.append(["revenue", 5_000_000, ""])
        ws.append(["currency", "USD", ""])
        ws.append(["company_name", "FileCo", ""])
        ws.append(["period", "Q1 2026", ""])
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        data = ingest_excel(str(path))
        assert data["revenue"] == 5_000_000.0


# ── ingest_file and ingest_bytes ──────────────────────────────────────────────

class TestIngestFile:
    def test_csv_by_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("revenue,5000000\ncurrency,USD\ncompany_name,C\nperiod,Q1 2026\n")
        data = ingest_file(str(f))
        assert data["revenue"] == 5_000_000.0

    def test_unsupported_extension_raises(self, tmp_path):
        f = tmp_path / "data.json"
        f.write_text("{}")
        with pytest.raises(ValueError, match="Unsupported"):
            ingest_file(str(f))

    def test_ingest_bytes_csv(self):
        raw = b"revenue,5000000\ncurrency,USD\n"
        data = ingest_bytes(raw, "upload.csv")
        assert data["revenue"] == 5_000_000.0

    def test_ingest_bytes_excel(self):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Field", "Value"])
        ws.append(["revenue", 7_000_000])
        ws.append(["currency", "USD"])
        buf = io.BytesIO()
        wb.save(buf)
        data = ingest_bytes(buf.getvalue(), "upload.xlsx")
        assert data["revenue"] == 7_000_000.0

    def test_ingest_bytes_unsupported(self):
        with pytest.raises(ValueError, match="Unsupported"):
            ingest_bytes(b"data", "file.pdf")


# ── Integration: ingest → FinancialDataSchema ─────────────────────────────────

class TestIngestToSchema:
    def test_csv_round_trip_to_schema(self):
        from backend.schemas.financial import FinancialDataSchema
        data = ingest_csv(io.StringIO(
            "company_name,Acme Corp\nperiod,Q1 2026\ncurrency,USD\n"
            "revenue,12500000\ncogs,5225000\ntotal_assets,45000000\ntotal_equity,28000000\n"
        ))
        schema = FinancialDataSchema(**{
            k: v for k, v in data.items()
            if k not in ("actuals", "budget", "historical_revenue", "line_items")
        })
        assert schema.company_name == "Acme Corp"
        assert schema.revenue == 12_500_000.0

    def test_excel_round_trip_to_schema(self):
        import importlib.util
        if importlib.util.find_spec("openpyxl") is None:
            pytest.skip("openpyxl not installed")
        from backend.schemas.financial import FinancialDataSchema
        buf = self._make_excel_buf([
            ["company_name", "ExcelCo", ""],
            ["period", "Q1 2026", ""],
            ["currency", "USD", ""],
            ["revenue", 10_000_000, ""],
        ])
        data = ingest_excel(buf)
        schema = FinancialDataSchema(**{
            k: v for k, v in data.items()
            if k not in ("actuals", "budget", "historical_revenue", "line_items")
        })
        assert schema.revenue == 10_000_000.0

    def _make_excel_buf(self, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
