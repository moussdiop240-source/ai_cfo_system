"""
Generate CFO Excel models:
1. Scenario Planning Model (Assumptions + Revenue + Sensitivity)
2. 3-Statement Financial Model (Revenue + Costs + HC + P&L + Cash Flow)

Color coding:
  Blue  (0,0,255)  : User inputs (hardcoded)
  Black (0,0,0)    : Formulas / calculated
  Green (0,128,0)  : Cross-sheet links
  Red   (255,0,0)  : External links
"""
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Color constants ────────────────────────────────────────────────────────
BLUE_INPUT  = "0000FF"
BLACK_CALC  = "000000"
GREEN_LINK  = "008000"
RED_EXT     = "FF0000"

HEADER_FILL   = PatternFill("solid", fgColor="1E293B")
SUBHEADER_FILL = PatternFill("solid", fgColor="334155")
INPUT_FILL    = PatternFill("solid", fgColor="EFF6FF")   # Light blue
FORMULA_FILL  = PatternFill("solid", fgColor="F8FAFC")   # Off-white
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
INPUT_FONT    = Font(color=BLUE_INPUT, bold=False, size=10)
CALC_FONT     = Font(color=BLACK_CALC, bold=False, size=10)
LINK_FONT     = Font(color=GREEN_LINK, bold=False, size=10)


def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def pct_fmt(): return "0.0%"
def usd_fmt(): return '$#,##0;($#,##0);"-"'
def usd_neg(): return '$#,##0;[Red]($#,##0);"-"'


def set_header(ws, row, col, text, span=1, bold=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT if bold else Font(bold=True, size=10)
    cell.fill = HEADER_FILL if bold else SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    return cell


def set_input(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border()
    return cell


def set_formula(ws, row, col, formula, color=BLACK_CALC, link=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = LINK_FONT if link else Font(color=color, size=10)
    cell.fill = FORMULA_FILL
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border()
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# SCENARIO PLANNING MODEL
# ═══════════════════════════════════════════════════════════════════════════════

def build_scenario_model(
    company: str = "Acme Corp",
    base_revenue: float = 10_000_000,
    products: list = None,
    output_path: str = "CFO_Scenario_Model.xlsx",
) -> str:
    if products is None:
        products = ["Product A", "Product B", "Product C", "Product D"]

    wb = Workbook()
    wb.remove(wb.active)

    _build_assumptions_sheet(wb, products, company)
    _build_product_revenue_sheet(wb, products, base_revenue)
    _build_sensitivity_sheet(wb, products, base_revenue)

    wb.save(output_path)
    return output_path


def _build_assumptions_sheet(wb, products, company):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = f"{company} — Scenario Planning Assumptions"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    note = ws["A2"]
    note.value = "Blue cells = change these. Black cells = calculated. Do not modify black cells."
    note.font = Font(italic=True, color="6B7280", size=9)

    set_header(ws, 4, 1, "Growth Rate Assumptions")
    set_header(ws, 4, 2, "BEAR Case")
    set_header(ws, 4, 3, "BASE Case")
    set_header(ws, 4, 4, "BULL Case")

    for i, prod in enumerate(products, 5):
        ws.cell(row=i, column=1, value=prod).font = Font(bold=True, size=10)
        set_input(ws, i, 2, -0.05 if i == 5 else -0.03)
        set_input(ws, i, 3, 0.10 if i == 5 else 0.08)
        set_input(ws, i, 4, 0.25 if i == 5 else 0.20)
        for c in [2, 3, 4]:
            ws.cell(row=i, column=c).number_format = pct_fmt()


def _build_product_revenue_sheet(wb, products, base_rev):
    ws = wb.create_sheet("Product Revenue")
    ws.column_dimensions["A"].width = 24
    years = [2024, 2025, 2026, 2027, 2028, 2029]
    scenarios = ["Bear", "Base", "Bull"]

    set_header(ws, 1, 1, "Product Revenue Model")
    col_offset = 2
    for j, (yr, sc) in enumerate([(y, s) for s in scenarios for y in years]):
        set_header(ws, 1, col_offset + j, f"{yr}\n{sc}", bold=False)
        ws.column_dimensions[get_column_letter(col_offset + j)].width = 12

    for i, prod in enumerate(products):
        row = 3 + i * 3
        ws.cell(row=row, column=1, value=prod).font = Font(bold=True, size=10)

        for j, (yr, sc_idx) in enumerate([(y, s) for s in range(3) for y in range(len(years))]):
            yr_val = years[yr]
            scenario = ["Bear", "Base", "Bull"][sc_idx]
            col = col_offset + j

            if yr == 0:  # 2024A — input
                val = base_rev / len(products)
                c = set_input(ws, row + 1, col, val)
                c.number_format = usd_fmt()
            else:
                prev_col = get_column_letter(col - 1)
                assump_row = 4 + i
                assump_col = get_column_letter(2 + sc_idx)
                formula = f"={prev_col}{row+1}*(1+Assumptions!{assump_col}{assump_row})"
                c = set_formula(ws, row + 1, col, formula)
                c.number_format = usd_fmt()

    # Total row
    total_row = 3 + len(products) * 3 + 1
    ws.cell(row=total_row, column=1, value="TOTAL REVENUE").font = Font(bold=True, size=10)
    for j in range(len(years) * 3):
        col = col_offset + j
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
        c = set_formula(ws, total_row, col, formula, BLACK_CALC)
        c.font = Font(bold=True, size=10)
        c.number_format = usd_fmt()


def _build_sensitivity_sheet(wb, products, base_rev):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.column_dimensions["A"].width = 20

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Revenue Sensitivity — Growth Rate × Gross Margin"
    ws["A1"].font = Font(bold=True, size=13)

    growth_rates = [-0.10, -0.05, 0.00, 0.05, 0.10, 0.15, 0.20, 0.25]
    margins      = [0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50]

    set_header(ws, 3, 1, "Revenue→EBITDA")
    for j, g in enumerate(growth_rates):
        c = ws.cell(row=3, column=j + 2, value=f"{g*100:.0f}%")
        c.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(j + 2)].width = 12

    for i, margin in enumerate(margins):
        ws.cell(row=4 + i, column=1, value=f"GM {margin*100:.0f}%").font = Font(bold=True)
        for j, growth in enumerate(growth_rates):
            value = round(base_rev * (1 + growth) * margin, 0)
            c = ws.cell(row=4 + i, column=j + 2, value=value)
            c.number_format = usd_fmt()

    # Conditional formatting: Green >$5M, Amber $2.5-5M, Red <$2.5M
    from openpyxl.formatting.rule import CellIsRule
    red    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    green  = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    rng = f"B4:{get_column_letter(1 + len(growth_rates))}{3 + len(margins)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=[str(base_rev * 0.5)], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(base_rev * 0.25), str(base_rev * 0.5)], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=[str(base_rev * 0.25)], fill=red))


# ═══════════════════════════════════════════════════════════════════════════════
# 3-STATEMENT FINANCIAL MODEL
# ═══════════════════════════════════════════════════════════════════════════════

def build_three_statement_model(
    company: str = "Acme Corp",
    base_revenue: float = 10_000_000,
    output_path: str = "CFO_3Statement_Model.xlsx",
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    products  = ["Product A", "Product B", "Product C", "Product D"]
    depts     = ["Engineering", "Sales", "Finance", "Customer Success", "R&D"]
    years     = [2024, 2025, 2026, 2027, 2028, 2029]
    actuals   = ["2024A"]
    estimates = [f"{y}E" for y in years[1:]]
    all_cols  = actuals + estimates

    _build_revenue_sheet(wb, products, years, all_cols, base_revenue)
    _build_cogs_sheet(wb, products, years, all_cols)
    _build_hc_sheet(wb, depts, years, all_cols)
    _build_pl_sheet(wb, years, all_cols, company)
    _build_cashflow_sheet(wb, years, all_cols)

    wb.save(output_path)
    return output_path


def _build_revenue_sheet(wb, products, years, all_cols, base_rev):
    ws = wb.create_sheet("1. Product Revenue")
    ws.column_dimensions["A"].width = 24

    set_header(ws, 1, 1, "Product Revenue")
    for j, col in enumerate(all_cols):
        set_header(ws, 1, j + 2, col, bold=False)
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    row = 3
    for prod in products:
        # Growth assumption row (blue)
        ws.cell(row=row, column=1, value=f"{prod} — Growth %").font = Font(italic=True, color="6B7280")
        set_input(ws, row, 2, None)  # 2024A — no growth assumption
        for j in range(1, len(all_cols)):
            c = set_input(ws, row, j + 2, 0.10)
            c.number_format = pct_fmt()

        # Revenue row (formula)
        ws.cell(row=row + 1, column=1, value=prod).font = Font(bold=True)
        c = set_input(ws, row + 1, 2, base_rev / len(products))
        c.number_format = usd_fmt()
        for j in range(1, len(all_cols)):
            prev_col = get_column_letter(j + 1)
            growth_col = get_column_letter(j + 2)
            formula = f"={prev_col}{row+1}*(1+{growth_col}{row})"
            c = set_formula(ws, row + 1, j + 2, formula)
            c.number_format = usd_fmt()

        row += 3

    # Total
    ws.cell(row=row, column=1, value="TOTAL REVENUE").font = Font(bold=True, size=11)
    for j, _ in enumerate(all_cols):
        col_letter = get_column_letter(j + 2)
        formula = f"=SUM({col_letter}4:{col_letter}{row-1})"
        c = set_formula(ws, row, j + 2, formula, BLACK_CALC)
        c.font = Font(bold=True, size=11)
        c.number_format = usd_fmt()

    # Embedded bar chart
    chart = BarChart()
    chart.title = "Total Revenue Forecast"
    chart.type = "col"
    chart.style = 10
    chart.grouping = "clustered"
    chart.y_axis.title = "Revenue ($)"
    data = Reference(ws, min_col=2, max_col=len(all_cols)+1, min_row=row, max_row=row)
    cats = Reference(ws, min_col=2, max_col=len(all_cols)+1, min_row=1, max_row=1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 12
    ws.add_chart(chart, f"A{row + 3}")


def _build_cogs_sheet(wb, products, years, all_cols):
    ws = wb.create_sheet("2. Cost of Product")
    ws.column_dimensions["A"].width = 28

    set_header(ws, 1, 1, "Cost of Product (COGS)")
    for j, col in enumerate(all_cols):
        set_header(ws, 1, j + 2, col, bold=False)
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    row = 3
    for prod in products:
        ws.cell(row=row, column=1, value=f"{prod} — COGS %").font = Font(italic=True, color="6B7280")
        for j in range(len(all_cols)):
            c = set_input(ws, row, j + 2, 0.45)
            c.number_format = pct_fmt()

        ws.cell(row=row + 1, column=1, value=f"{prod} COGS").font = Font(bold=True)
        for j in range(len(all_cols)):
            col_letter = get_column_letter(j + 2)
            rev_formula = f"='1. Product Revenue'!{col_letter}{(list(products).index(prod))*3 + 4}"
            formula = f"={rev_formula}*{col_letter}{row}"
            c = set_formula(ws, row + 1, j + 2, formula, link=True)
            c.number_format = usd_fmt()
        row += 3

    # Totals
    ws.cell(row=row, column=1, value="TOTAL COGS").font = Font(bold=True, size=11)
    ws.cell(row=row+1, column=1, value="GROSS PROFIT").font = Font(bold=True, size=11)
    ws.cell(row=row+2, column=1, value="GROSS MARGIN %").font = Font(bold=True, size=11)

    for j in range(len(all_cols)):
        col_letter = get_column_letter(j + 2)
        set_formula(ws, row, j + 2, f"=SUM({col_letter}4:{col_letter}{row-1})")
        rev_row_ref = f"='1. Product Revenue'!{col_letter}{(len(products))*3 + 4}"
        gp_formula = f"={rev_row_ref}-{col_letter}{row}"
        c = set_formula(ws, row+1, j + 2, gp_formula)
        c.number_format = usd_fmt()
        gm_formula = f"={col_letter}{row+1}/{rev_row_ref}"
        c = set_formula(ws, row+2, j + 2, gm_formula)
        c.number_format = pct_fmt()


def _build_hc_sheet(wb, depts, years, all_cols):
    ws = wb.create_sheet("3. Headcount Planning")
    ws.column_dimensions["A"].width = 28

    set_header(ws, 1, 1, "Headcount Planning")
    for j, col in enumerate(all_cols):
        set_header(ws, 1, j + 2, col, bold=False)
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    row = 3
    for dept in depts:
        ws.cell(row=row, column=1, value=f"{dept} — HC Growth").font = Font(italic=True, color="6B7280")
        set_input(ws, row, 2, None)
        for j in range(1, len(all_cols)):
            c = set_input(ws, row, j + 2, 0.15)
            c.number_format = pct_fmt()

        ws.cell(row=row+1, column=1, value=f"{dept} — Avg Salary ($K)").font = Font(italic=True, color="6B7280")
        for j in range(len(all_cols)):
            c = set_input(ws, row+1, j + 2, 120 + j * 4)  # 4% annual escalation
            c.number_format = usd_fmt()

        ws.cell(row=row+2, column=1, value=f"{dept} — HC").font = Font(bold=True)
        c = set_input(ws, row+2, 2, 10)
        for j in range(1, len(all_cols)):
            prev_col = get_column_letter(j + 1)
            growth_col = get_column_letter(j + 2)
            formula = f"=ROUND({prev_col}{row+2}*(1+{growth_col}{row}),0)"
            c = set_formula(ws, row+2, j + 2, formula)

        ws.cell(row=row+3, column=1, value=f"{dept} — Personnel Cost ($K)").font = Font(bold=False)
        for j in range(len(all_cols)):
            col_letter = get_column_letter(j + 2)
            formula = f"={col_letter}{row+2}*{col_letter}{row+1}/1000"
            c = set_formula(ws, row+3, j + 2, formula)
            c.number_format = usd_fmt()

        row += 5


def _build_pl_sheet(wb, years, all_cols, company):
    ws = wb.create_sheet("4. P&L Summary")
    ws.column_dimensions["A"].width = 32

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"{company} — Income Statement"
    ws["A1"].font = Font(bold=True, size=14)

    set_header(ws, 2, 1, "P&L Summary ($K)")
    for j, col in enumerate(all_cols):
        set_header(ws, 2, j + 2, col, bold=False)
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    labels = [
        ("Revenue", "green"), ("COGS", "green"),
        ("Gross Profit", "calc"), ("Gross Margin %", "pct"),
        ("", ""), ("SG&A", "input"), ("R&D", "input"),
        ("D&A", "input"), ("EBITDA", "calc"), ("EBITDA Margin %", "pct"),
        ("", ""), ("Interest Expense", "input"), ("Pre-Tax Income", "calc"),
        ("Tax Provision", "input"), ("Net Income", "calc"), ("Net Margin %", "pct"),
        ("", ""), ("Basic EPS", "calc"), ("Diluted EPS", "calc"),
    ]

    for i, (label, typ) in enumerate(labels, 4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=(typ in ("calc", "pct")), size=10)
        for j in range(len(all_cols)):
            col_letter = get_column_letter(j + 2)
            if typ == "input":
                c = set_input(ws, i, j + 2, 0)
                c.number_format = usd_fmt()
            elif typ == "green":
                c = set_formula(ws, i, j + 2, f"='1. Product Revenue'!{col_letter}20", link=True)
                c.number_format = usd_fmt()
            elif label == "Gross Profit":
                c = set_formula(ws, i, j + 2, f"={col_letter}4-{col_letter}5")
                c.number_format = usd_fmt()


def _build_cashflow_sheet(wb, years, all_cols):
    ws = wb.create_sheet("5. Cash Flow")
    ws.column_dimensions["A"].width = 32

    set_header(ws, 1, 1, "Cash Flow Statement")
    for j, col in enumerate(all_cols):
        set_header(ws, 1, j + 2, col, bold=False)
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    labels = [
        ("Net Income", "link"), ("+ D&A", "link"), ("WC Changes", "input"),
        ("Cash from Operations", "calc"), ("", ""),
        ("Capex", "input"), ("Cash from Investing", "calc"), ("", ""),
        ("Free Cash Flow", "calc"),
    ]

    for i, (label, typ) in enumerate(labels, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=(typ == "calc"), size=10)


# ── Entry points ────────────────────────────────────────────────────────────

def generate_all(output_dir: str = ".", company: str = "Acme Corp", base_revenue: float = 10_000_000):
    os.makedirs(output_dir, exist_ok=True)

    scenario = build_scenario_model(
        company=company,
        base_revenue=base_revenue,
        output_path=os.path.join(output_dir, "CFO_Scenario_Model.xlsx"),
    )

    three_stmt = build_three_statement_model(
        company=company,
        base_revenue=base_revenue,
        output_path=os.path.join(output_dir, "CFO_3Statement_Model.xlsx"),
    )

    print(f"Generated:\n  {scenario}\n  {three_stmt}")
    return scenario, three_stmt


if __name__ == "__main__":
    generate_all(output_dir="./output", company="Acme Corp", base_revenue=25_000_000)
